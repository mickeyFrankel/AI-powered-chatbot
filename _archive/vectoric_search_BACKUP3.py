import os
from dotenv import load_dotenv
load_dotenv()

# Suppress ChromaDB telemetry warnings
os.environ['ANONYMIZED_TELEMETRY'] = 'False'
import sys
import io
import logging

# Suppress ChromaDB logging  
logging.getLogger('chromadb').setLevel(logging.ERROR)
logging.getLogger('chromadb.telemetry').setLevel(logging.CRITICAL)

import re, unicodedata
import pandas as pd
import numpy as np
import chromadb

# Patch chromadb's telemetry to suppress errors
try:
    import chromadb.telemetry.posthog as posthog
    original_capture = posthog.Posthog.capture
    posthog.Posthog.capture = lambda *args, **kwargs: None
except:
    pass
import json
from chromadb.config import Settings
from sentence_transformers import SentenceTransformer

# LangChain imports
try:
    from langchain_core.tools import tool
except ImportError:
    from langchain.tools import tool

from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate, MessagesPlaceholder
from langchain_core.messages import AIMessage, HumanMessage

try:
    from langchain.agents import create_openai_functions_agent, AgentExecutor
except ImportError:
    try:
        from langchain.agents.openai_functions_agent.base import create_openai_functions_agent
        from langchain.agents.agent import AgentExecutor
    except ImportError:
        # Fallback for older versions
        create_openai_functions_agent = None
        AgentExecutor = None

import docx
from openpyxl import load_workbook
import PyPDF2
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from collections import Counter
import warnings

from sklearn.metrics.pairwise import cosine_similarity, euclidean_distances

# Fuzzy matching for typo tolerance
try:
    from rapidfuzz import fuzz, process
    HAS_FUZZY = True
except ImportError:
    HAS_FUZZY = False
    print("⚠️  rapidfuzz not installed. Fuzzy search disabled.")
    print("   Install with: pip install rapidfuzz")

warnings.filterwarnings('ignore')


class VectorDBQASystem:
    """Base VectorDB system - keep unchanged"""
    def __init__(self, persist_directory: str = "./chroma_db", model_name: str = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"):
        """
        Initialize the VectorDB Q&A System with Hebrew support
        
        Args:
            persist_directory: Directory to persist ChromaDB data
            model_name: Sentence transformer model (supports Hebrew)
        """
        self.persist_directory = persist_directory
        self.model_name = model_name
        
        # Initialize embedding model (supports Hebrew)
        print(f"Loading embedding model: {model_name}")
        self.embedding_model = SentenceTransformer(model_name)
        
        # Initialize ChromaDB client
        self.client = chromadb.PersistentClient(
            path=persist_directory,
            settings=Settings(
                anonymized_telemetry=False,
                allow_reset=True
            )
        )
        
        # Create or get collection
        self.collection_name = "documents"
        try:
            self.collection = self.client.get_collection(self.collection_name)
            print(f"Loaded existing collection: {self.collection_name}")
        except:
            self.collection = self.client.create_collection(
                name=self.collection_name,
                metadata={"hnsw:space": "cosine"}
            )
            print(f"Created new collection: {self.collection_name}")
    
    def reset_database(self):
        """Reset database by deleting and recreating collection"""
        if hasattr(self, 'chat_history'):
            self.chat_history = []
        
        try:
            self.client.delete_collection(self.collection_name)
        except:
            pass  # Collection might not exist
        
        self.collection = self.client.create_collection(
            name=self.collection_name,
            metadata={"hnsw:space": "cosine"}
        )
    
    def _preprocess_dataframe(self, df: pd.DataFrame, source_name: str = "") -> pd.DataFrame:
        """Clean and optimize DataFrame before ingestion"""
        print(f"\n🧹 Preprocessing {source_name}...")
        print(f"   Initial: {len(df)} rows × {len(df.columns)} columns")
        
        original_cols = len(df.columns)
        original_rows = len(df)
        
        # 1. Remove completely empty columns (100% null)
        empty_cols = df.columns[df.isnull().all()].tolist()
        if empty_cols:
            df = df.drop(columns=empty_cols)
            print(f"   ✂️  Removed {len(empty_cols)} empty columns: {empty_cols[:3]}..." if len(empty_cols) > 3 else f"   ✂️  Removed {len(empty_cols)} empty columns: {empty_cols}")
        
        # 2. Remove sparse columns (>95% empty) BUT protect important fields
        sparse_threshold = 0.95
        important_patterns = ['phone', 'mobile', 'tel', 'email', 'mail', 'address', 'addr']
        sparse_cols = []
        
        for col in df.columns:
            null_ratio = df[col].isnull().sum() / len(df)
            if null_ratio > sparse_threshold:
                # Check if it's an important field that should be kept
                col_lower = col.lower()
                is_important = any(pattern in col_lower for pattern in important_patterns)
                if not is_important:
                    sparse_cols.append(col)
        
        if sparse_cols:
            df = df.drop(columns=sparse_cols)
            print(f"   ✂️  Removed {len(sparse_cols)} sparse columns (>95% empty)")
        
        # 3. Remove low-value metadata columns
        low_value_patterns = ['id', 'uuid', 'guid', 'key', 'index', 'row_num', 'created_at', 'updated_at', 'timestamp', 'date_added', 'last_modified']
        metadata_cols = []
        for col in df.columns:
            col_lower = col.lower()
            if any(pattern in col_lower for pattern in low_value_patterns):
                # Keep if it looks like a phone ("mobile_id" might be "mobile identifier")
                if 'phone' not in col_lower and 'mobile' not in col_lower and 'tel' not in col_lower:
                    metadata_cols.append(col)
        
        if metadata_cols:
            df = df.drop(columns=metadata_cols)
            print(f"   ✂️  Removed {len(metadata_cols)} metadata columns")
        
        # 4. Trim whitespace from all string columns
        for col in df.columns:
            if df[col].dtype == 'object':  # String columns
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        # 5. Remove duplicate rows (exact duplicates)
        duplicates = df.duplicated().sum()
        if duplicates > 0:
            df = df.drop_duplicates()
            print(f"   ✂️  Removed {duplicates} duplicate rows")
        
        # 6. Consolidate phone columns (prioritize 'value' over 'type/label')
        phone_cols = [col for col in df.columns if 'phone' in col.lower() or 'mobile' in col.lower() or 'tel' in col.lower()]
        if len(phone_cols) > 1:
            # Sort to prioritize columns with 'value', 'number', or numeric data
            value_cols = [col for col in phone_cols if 'value' in col.lower() or 'number' in col.lower()]
            type_cols = [col for col in phone_cols if 'type' in col.lower() or 'label' in col.lower()]
            other_cols = [col for col in phone_cols if col not in value_cols and col not in type_cols]
            
            # Prioritize: value columns > other columns > type columns (last)
            sorted_phone_cols = value_cols + other_cols + type_cols
            
            # Create single 'phone' column, filtering out non-numeric values
            def get_phone_value(row):
                for col in sorted_phone_cols:
                    val = row[col]
                    if pd.notna(val) and str(val).strip():
                        val_str = str(val).strip()
                        # Skip if it looks like a label (e.g., "Mobile", "נייד")
                        # Phone numbers must contain at least one digit
                        if any(c.isdigit() for c in val_str):
                            return val
                return None
            
            df['phone'] = df[sorted_phone_cols].apply(get_phone_value, axis=1)
            
            # Drop original phone columns
            df = df.drop(columns=phone_cols)
            print(f"   🔗 Consolidated {len(phone_cols)} phone columns into 'phone'")
        elif len(phone_cols) == 1:
            # Rename single phone column to 'phone'
            df = df.rename(columns={phone_cols[0]: 'phone'})
            print(f"   🔗 Renamed '{phone_cols[0]}' to 'phone'")
        
        # 7. Remove rows that are mostly empty (>90% null)
        row_null_threshold = 0.9
        rows_to_keep = []
        for idx, row in df.iterrows():
            null_ratio = row.isnull().sum() / len(row)
            if null_ratio < row_null_threshold:
                rows_to_keep.append(idx)
        
        removed_rows = len(df) - len(rows_to_keep)
        if removed_rows > 0:
            df = df.loc[rows_to_keep]
            print(f"   ✂️  Removed {removed_rows} rows with insufficient data")
        
        # 8. Reorder columns: put important ones first
        priority_cols = ['name', 'phone', 'email', 'address', 'company', 'title', 'notes']
        remaining_cols = [col for col in df.columns if col not in priority_cols]
        ordered_cols = [col for col in priority_cols if col in df.columns] + remaining_cols
        df = df[ordered_cols]
        
        cols_removed = original_cols - len(df.columns)
        rows_removed = original_rows - len(df)
        
        print(f"   ✅ Final: {len(df)} rows × {len(df.columns)} columns")
        print(f"   📊 Reduced by {cols_removed} columns and {rows_removed} rows")
        print(f"   💾 Data size reduction: ~{(cols_removed + rows_removed) / (original_cols + original_rows) * 100:.1f}%\n")
        
        return df
    
    def read_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Read CSV file and return list of documents with phone number fixing"""
        df = pd.read_csv(file_path, encoding='utf-8')
        
        # PREPROCESS: Clean and optimize
        df = self._preprocess_dataframe(df, source_name=Path(file_path).name)
        
        # FIX PHONE NUMBERS BEFORE PROCESSING
        for col in df.columns:
            if 'phone' in col.lower() or 'mobile' in col.lower() or 'tel' in col.lower():
                df[col] = df[col].apply(self._fix_phone_number)
        
        documents = []
        
        for idx, row in df.iterrows():
            text_parts = []
            metadata = {}
            
            for col in df.columns:
                if pd.notna(row[col]):
                    text_parts.append(f"{col}: {row[col]}")
                    metadata[col] = str(row[col])
            
            documents.append({
                'text': ' | '.join(text_parts),
                'metadata': metadata,
                'source': file_path,
                'row_id': idx
            })
        
        return documents
    
    def _fix_phone_number(self, value):
        """Convert scientific notation to proper phone format"""
        if pd.isna(value) or value == '':
            return value
        
        value_str = str(value).strip()
        
        # Detect scientific notation (9.73E+11, 5.42e+08)
        if 'e+' in value_str.lower() or 'e-' in value_str.lower():
            try:
                # Convert to integer
                num = int(float(value_str))
                result = str(num)
                
                # Israeli phone: 9-10 digits, add leading 0 if missing
                if len(result) in [9, 10]:
                    if not result.startswith('0'):
                        result = '0' + result
                
                print(f"Fixed phone: {value_str} → {result}")
                return result
            except (ValueError, OverflowError):
                return value
        
        return value
    
    def read_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Read Excel file and return list of documents with phone number fixing"""
        workbook = load_workbook(file_path, read_only=True)
        documents = []
        
        for sheet_name in workbook.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # PREPROCESS: Clean and optimize
            df = self._preprocess_dataframe(df, source_name=f"{Path(file_path).name}/{sheet_name}")
            
            # FIX PHONE NUMBERS BEFORE PROCESSING
            for col in df.columns:
                if 'phone' in col.lower() or 'mobile' in col.lower() or 'tel' in col.lower():
                    df[col] = df[col].apply(self._fix_phone_number)
            
            for idx, row in df.iterrows():
                text_parts = []
                metadata = {'sheet': sheet_name}
                
                for col in df.columns:
                    if pd.notna(row[col]):
                        text_parts.append(f"{col}: {row[col]}")
                        metadata[col] = str(row[col])
                
                documents.append({
                    'text': ' | '.join(text_parts),
                    'metadata': metadata,
                    'source': file_path,
                    'row_id': f"{sheet_name}_{idx}"
                })
        
        return documents
    
    def read_docx(self, file_path: str) -> List[Dict[str, Any]]:
        """Read DOCX file and return list of documents"""
        doc = docx.Document(file_path)
        documents = []
        
        for i, paragraph in enumerate(doc.paragraphs):
            if paragraph.text.strip():
                documents.append({
                    'text': paragraph.text.strip(),
                    'metadata': {'paragraph_id': i},
                    'source': file_path,
                    'row_id': f"para_{i}"
                })
        
        return documents
    
    def read_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """Read PDF file and return list of documents"""
        documents = []
        
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                if text.strip():
                    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
                    
                    for para_idx, paragraph in enumerate(paragraphs):
                        documents.append({
                            'text': paragraph,
                            'metadata': {
                                'page': page_num + 1,
                                'paragraph_id': para_idx
                            },
                            'source': file_path,
                            'row_id': f"page_{page_num}_para_{para_idx}"
                        })
        
        return documents
    
    def read_txt(self, file_path: str) -> List[Dict[str, Any]]:
        """Read TXT file and return list of documents"""
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
        documents = []
        
        for i, paragraph in enumerate(paragraphs):
            documents.append({
                'text': paragraph,
                'metadata': {'paragraph_id': i},
                'source': file_path,
                'row_id': f"para_{i}"
            })
        
        return documents
    
    def load_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Load file based on extension"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        extension = file_path.suffix.lower()
        
        print(f"Loading file: {file_path} (type: {extension})")
        
        if extension == '.csv':
            return self.read_csv(str(file_path))
        elif extension in ['.xlsx', '.xls']:
            return self.read_excel(str(file_path))
        elif extension == '.docx':
            return self.read_docx(str(file_path))
        elif extension == '.pdf':
            return self.read_pdf(str(file_path))
        elif extension == '.txt':
            return self.read_txt(str(file_path))
        else:
            raise ValueError(f"Unsupported file format: {extension}")
    
    def ingest_file(self, file_path: str) -> dict:
        """Load a file and add its documents to the database. Returns ingestion report."""
        documents = self.load_file(file_path)
        self.add_documents(documents)
        
        return {
            'documents_added': len(documents),
            'file': Path(file_path).name,
            'total_in_db': self.collection.count()
        }
    
    def add_documents(self, documents: List[Dict[str, Any]]):
        """Add documents to ChromaDB with deterministic IDs; skip existing rows."""
        if not documents:
            print("No documents to add.")
            return

        abs_paths = [str(Path(d['source']).resolve()) for d in documents]
        source_keys = [p.lower() for p in abs_paths]
        ids = [f"{sk}::{d['row_id']}" for sk, d in zip(source_keys, documents)]

        existing = set(self.collection.get(ids=ids).get('ids', []) or [])
        to_add_idx = [i for i, _id in enumerate(ids) if _id not in existing]
        if not to_add_idx:
            print("Nothing new to add (all IDs already present).")
            return

        new_docs = [documents[i] for i in to_add_idx]
        new_ids  = [ids[i]        for i in to_add_idx]

        texts = [doc['text'] for doc in new_docs]
        print(f"Generating embeddings for {len(texts)} new documents...")
        embeddings = self.embedding_model.encode(texts, show_progress_bar=True).tolist()

        metadatas = []
        for doc in new_docs:
            src_path = str(Path(doc['source']).resolve())
            src_name = Path(src_path).name
            src_key  = src_path.lower()

            md = (doc.get('metadata') or {}).copy()
            md.update({
                'source': src_path,
                'source_name': src_name,
                'source_key': src_key,
                'row_id': doc['row_id'],
            })

            stats = self._derive_name_fields(doc['text'], metadata=md)
            md.update(stats)
            metadatas.append(md)

        print(f"Adding to ChromaDB... ({len(new_ids)} new)")
        self.collection.add(
            embeddings=embeddings,
            documents=texts,
            metadatas=metadatas,
            ids=new_ids
        )
        print(f"Successfully added {len(new_ids)} new documents to the vector database!")

    def purge_duplicates(self):
        got = self.collection.get(include=["metadatas"])
        ids   = got.get("ids") or []
        metas = got.get("metadatas") or []
        if not ids or not metas or len(ids) != len(metas):
            print("⚠️ Could not read IDs/metadata reliably; skipping dedupe.")
            return
        def _basename(md):
            s = (md.get("source_name") or md.get("source") or "")
            try: return Path(s).name.lower()
            except Exception: return str(s).split("\\")[-1].split("/")[-1].lower()
        seen, to_delete = set(), []
        for _id, md in zip(ids, metas):
            key = (_basename(md), md.get("row_id"))
            if key in seen: to_delete.append(_id)
            else: seen.add(key)
        if to_delete:
            self.collection.delete(ids=to_delete)
            print(f"Deleted {len(to_delete)} duplicate records.")
        else:
            print("No duplicates found.")

    def _fuzzy_text_search(self, query: str, n_results: int = 100, threshold: int = 60) -> Dict[str, Any]:
        """Fuzzy text search - finds similar strings even with typos/variations"""
        from rapidfuzz import fuzz, process
        
        all_data = self.collection.get(include=["documents", "metadatas"])
        documents = all_data.get("documents", [])
        metadatas = all_data.get("metadatas", [])
        ids = all_data.get("ids", [])
        
        if not documents:
            return {'results': []}
        
        matches = []
        for i, (doc, meta, doc_id) in enumerate(zip(documents, metadatas, ids)):
            similarity = fuzz.partial_ratio(query.lower(), doc.lower())
            
            if similarity >= threshold:
                matches.append({
                    'id': doc_id,
                    'document': doc,
                    'metadata': meta,
                    'relevance': similarity / 100.0,
                    'similarity': similarity
                })
        
        matches.sort(key=lambda x: x['relevance'], reverse=True)
        
        return {'results': matches[:n_results]}
    
    def search(self, query: str, n_results: int = 5, similarity_metric: str = "cosine", 
               fuzzy_threshold: float = 80.0, auto_correct: bool = True, 
               hybrid: bool = True) -> Dict[str, Any]:
        """Hybrid search: tries text matching first, then semantic search."""
        original_query = query
        corrected_terms = []
        suggestions = []
        search_method = "hybrid"
        
        if HAS_FUZZY and auto_correct:
            corrected_query, corrections = self._fuzzy_correct_query(query, fuzzy_threshold)
            if corrections:
                corrected_terms = corrections
                query = corrected_query
        elif HAS_FUZZY:
            suggestions = self._get_fuzzy_suggestions(query, fuzzy_threshold)
            if suggestions:
                print(f"\n💡 Did you mean: {', '.join([s[0] for s in suggestions[:3]])}?")
        
        text_results = None
        if hybrid and len(query.strip()) >= 2:
            text_results = self._fuzzy_text_search(query, n_results=n_results, threshold=60)
            
            if text_results['results']:
                search_method = "fuzzy_text"
                
                formatted_results = {
                    'query': query,
                    'original_query': original_query,
                    'corrected': corrected_terms,
                    'suggestions': suggestions,
                    'search_method': search_method,
                    'results': []
                }
                
                for result in text_results['results']:
                    formatted_results['results'].append({
                        'id': result['id'],
                        'document': result['document'],
                        'metadata': result['metadata'],
                        'similarity_score': result['relevance'],
                        'match_type': 'text'
                    })
                
                return formatted_results
        
        search_method = "semantic"
        
        query_embedding = self.embedding_model.encode([query])
        
        results = self.collection.query(
            query_embeddings=query_embedding.tolist(),
            n_results=n_results
        )
        
        formatted_results = {
            'query': query,
            'original_query': original_query,
            'corrected': corrected_terms,
            'suggestions': suggestions,
            'search_method': search_method,
            'results': []
        }
        
        if results['documents'] and results['documents'][0]:
            for i in range(len(results['documents'][0])):
                result = {
                    'id': results['ids'][0][i],
                    'document': results['documents'][0][i],
                    'metadata': results['metadatas'][0][i],
                    'distance': results['distances'][0][i],
                    'similarity_score': 1 - results['distances'][0][i],
                    'match_type': 'semantic'
                }
                formatted_results['results'].append(result)
        
        return formatted_results
    
    def search_full_text(self, keyword: str, limit: int = 50) -> list[dict]:
        """Search for keyword across ALL document text with highlighted context"""
        # Common Hebrew profession synonyms
        synonyms = {
            'עורך דין': ['עורך דין', 'עו"ד', 'עוד', 'lawyer', 'attorney'],
            'עו"ד': ['עורך דין', 'עו"ד', 'עוד', 'lawyer', 'attorney'],
            'רופא': ['רופא', 'ד"ר', 'דר', 'doctor', 'dr'],
            'דוקטור': ['רופא', 'ד"ר', 'דר', 'doctor', 'dr', 'דוקטור'],
        }
        
        # Check if keyword has synonyms
        search_terms = synonyms.get(keyword.strip(), [keyword])
        
        all_data = self.collection.get(include=["documents", "metadatas"])
        documents = all_data.get("documents", [])
        metadatas = all_data.get("metadatas", [])
        
        matches = []
        seen_names = set()  # Avoid duplicates
        
        for search_term in search_terms:
            keyword_norm = self._normalize(search_term).upper()
            
            for doc, meta in zip(documents, metadatas):
                name = meta.get('name', 'Unknown')
                if name in seen_names:
                    continue
                    
                doc_norm = self._normalize(doc).upper()
                if keyword_norm in doc_norm:
                    seen_names.add(name)
                    # Extract context around the keyword
                    context = self._extract_keyword_context(doc, search_term)
                    
                    matches.append({
                        'document': doc,
                        'metadata': meta,
                        'name': name,
                        'keyword_context': context,
                        'matched_term': search_term  # Which variant matched
                    })
                    
                    if len(matches) >= limit:
                        return matches
        
        return matches
    
    def _extract_keyword_context(self, text: str, keyword: str, window: int = 100) -> str:
        """Extract snippet showing where keyword appears in text"""
        keyword_norm = self._normalize(keyword).upper()
        text_norm = self._normalize(text).upper()
        
        # Find keyword position
        pos = text_norm.find(keyword_norm)
        if pos == -1:
            return text[:200]  # Fallback: show beginning
        
        # Extract window around keyword
        start = max(0, pos - window)
        end = min(len(text), pos + len(keyword) + window)
        
        snippet = text[start:end].strip()
        
        # Add ellipsis if truncated
        if start > 0:
            snippet = "..." + snippet
        if end < len(text):
            snippet = snippet + "..."
        
        return snippet
    
    def names_containing(self, substring: str, limit: int = 200) -> list[str]:
        """Find all names containing the substring (case-insensitive)"""
        s = self._normalize(substring)
        names = self.get_all_names()
        out = [nm for nm in names if s.upper() in nm.upper()]
        return sorted(set(out), key=str.upper)[:limit]
    
    def get_collection_stats(self) -> Dict[str, Any]:
        """Get statistics about the collection"""
        count = self.collection.count()
        return {
            'document_count': count,
            'collection_name': self.collection_name,
            'embedding_model': self.model_name
        }
    
    def list_by_prefix(self, letter: str) -> list[str]:
        letter = letter.upper().strip()
        all_docs = self.collection.get(include=["metadatas"])
        names = []
        for md in all_docs.get("metadatas", []):
            name = (md.get("name") or "").strip()
            if not name:
                continue
            # Get first alphabetic character
            first_char = None
            for ch in name:
                if ch.isalpha():
                    first_char = ch.upper()
                    break
            if first_char and first_char == letter:
                names.append(name)
        return sorted(set(names), key=str.upper)

    def _build_vocabulary(self) -> set:
        """Build vocabulary from all documents in the database."""
        all_docs = self.collection.get(include=["documents"])
        documents = all_docs.get("documents", [])
        
        vocabulary = set()
        for doc in documents:
            words = re.findall(r'\b[a-zA-Z0-9]+\b', doc.lower())
            vocabulary.update(words)
        
        return vocabulary
    
    def _fuzzy_correct_query(self, query: str, threshold: float = 80.0) -> Tuple[str, List[Tuple[str, str]]]:
        """Correct typos in query using fuzzy matching against database vocabulary."""
        if not HAS_FUZZY:
            return query, []
        
        vocabulary = self._build_vocabulary()
        
        if not vocabulary:
            return query, []
        
        query_words = query.split()
        corrected_words = []
        corrections = []
        
        for word in query_words:
            word_lower = word.lower()
            
            if word_lower in vocabulary:
                corrected_words.append(word)
                continue
            
            matches = process.extract(
                word_lower,
                vocabulary,
                scorer=fuzz.ratio,
                limit=1
            )
            
            if matches and matches[0][1] >= threshold:
                best_match = matches[0][0]
                if word[0].isupper():
                    best_match = best_match.capitalize()
                corrected_words.append(best_match)
                corrections.append((word, best_match))
            else:
                corrected_words.append(word)
        
        corrected_query = " ".join(corrected_words)
        return corrected_query, corrections
    
    def _get_fuzzy_suggestions(self, query: str, threshold: float = 70.0) -> List[Tuple[str, float]]:
        """Get fuzzy match suggestions for query terms without auto-correcting."""
        if not HAS_FUZZY:
            return []
        
        vocabulary = self._build_vocabulary()
        if not vocabulary:
            return []
        
        all_suggestions = []
        for word in query.split():
            word_lower = word.lower()
            if word_lower not in vocabulary:
                matches = process.extract(
                    word_lower,
                    vocabulary,
                    scorer=fuzz.ratio,
                    limit=3
                )
                for match_word, score in matches:
                    if score >= threshold:
                        all_suggestions.append((match_word, score))
        
        return sorted(all_suggestions, key=lambda x: x[1], reverse=True)[:5]
    
    def _normalize(self, s: str) -> str:
        """Normalize text for search - handles Hebrew Unicode properly"""
        if not s:
            return ""
        
        # First apply NFKC normalization
        s = unicodedata.normalize("NFKC", s)
        
        # Remove Hebrew vowel points (nikud) and cantillation marks
        # Range U+0591-U+05C7 covers most Hebrew diacritics
        s = re.sub(r'[\u0591-\u05C7]', '', s)
        
        # Remove zero-width characters and soft hyphens
        s = re.sub(r'[\u200B-\u200D\uFEFF\u00AD]', '', s)
        
        # Normalize whitespace
        s = re.sub(r'\s+', ' ', s)
        
        return s.strip()
    
    def _derive_name_fields(self, text: str, metadata: Optional[dict] = None) -> dict:
        name = ""
        md = metadata or {}
        
        # Try First + Last Name (Google Contacts CSV)
        first = md.get('First Name', '').strip()
        last = md.get('Last Name', '').strip()
        
        if first or last:
            name = ' '.join([p for p in [first, last] if p])
        
        # Fallback to Organization Name
        if not name:
            name = md.get('Organization Name', '').strip()
        
        # Last resort: extract from text
        if not name:
            preferred_keys = ['name', 'title', 'industry_name']
            for k in preferred_keys:
                v = md.get(k)
                if v and str(v).strip():
                    name = str(v).strip()
                    break
        
        if not name:
            t = self._normalize(text)
            first_chunk = t.split(" | ")[0]
            if ":" in first_chunk:
                maybe = first_chunk.split(":", 1)[1].strip()
                if maybe:
                    name = maybe
            if not name:
                name = first_chunk.strip()
        
        first_char = ""
        for ch in name:
            if ch.isalpha():
                first_char = ch.upper() if ch.isascii() else ch
                break
        
        return {
            "name": name,
            "first_char": first_char,
            "name_len": len(name),
        }
    def _get_all_metadatas(self) -> List[dict]:
        out = self.collection.get(include=["metadatas"])
        return out.get("metadatas", []) or []

    def get_all_names(self) -> List[str]:
        names = []
        for md in self._get_all_metadatas():
            name = self._normalize(md.get("name") or "")
            if name:
                names.append(name)
        return sorted(set(names), key=str.upper)

    def letter_histogram(self) -> Dict[str, int]:
        hist = {}
        for md in self._get_all_metadatas():
            ch = md.get("first_char") or ""
            if ch:
                hist[ch] = hist.get(ch, 0) + 1
        return dict(sorted(hist.items(), key=lambda kv: kv[0]))

    def length_histogram(self) -> Dict[int, int]:
        hist = {}
        for md in self._get_all_metadatas():
            n = int(md.get("name_len") or 0)
            hist[n] = hist.get(n, 0) + 1
        return dict(sorted(hist.items()))

    def first_n_by_prefix(self, prefix: str, n: int = 5) -> List[str]:
        p = self._normalize(prefix)
        names = self.get_all_names()
        return sorted([nm for nm in names if nm.upper().startswith(p.upper())], key=str.upper)[:n]

    def count_by_prefix(self, prefix: str) -> int:
        p = self._normalize(prefix)
        return sum(1 for md in self._get_all_metadatas()
                if (md.get("name") or "").upper().startswith(p.upper()))
    
    def _clean_for_len(self, s: str) -> str:
        return re.sub(r"[^0-9A-Za-z\u0590-\u05FF]", "", s)

    def _measure_len(self, name: str, count_mode: str = "chars") -> int:
        if count_mode == "letters":
            return len(self._clean_for_len(name))
        if count_mode == "words":
            return len([w for w in re.split(r"\s+", name.strip()) if w])
        return len(name)
    
    def names_by_length(self, length: int, limit: int = 200, count_mode="chars") -> List[str]:
        names = self.get_all_names()
        out = [nm for nm in names if self._measure_len(nm, count_mode) == length]
        return sorted(out, key=str.upper)[:limit]
    
    def names_by_prefix_and_length(self, prefix: str, length: int, limit: int = 200, count_mode="chars") -> list[str]:
        p = self._normalize(prefix)
        L = int(length)
        names = self.get_all_names()
        out = [nm for nm in names if nm.upper().startswith(p.upper()) and self._measure_len(nm, count_mode) == L]
        return sorted(set(out), key=str.upper)[:limit]

    def list_sources(self) -> Dict[str, int]:
        """Return {source_name: count} with canonical grouping."""
        out = self.collection.get(include=['metadatas'])
        metas = out.get('metadatas', []) or []
        counts = Counter()
        for md in metas:
            name = md.get('source_name') or md.get('source') or 'UNKNOWN'
            key  = md.get('source_key') or (md.get('source') or '').lower()
            counts[(key, name)] += 1
        display = Counter()
        for (_, name), c in counts.items():
            display[name] += c
        return dict(display)

    def count_sources(self) -> int:
        """Number of distinct canonical sources (by source_key)."""
        out = self.collection.get(include=['metadatas'])
        metas = out.get('metadatas', []) or []
        keys = {md.get('source_key') or (md.get('source') or '').lower() for md in metas if (md.get('source') or '')}
        return len(keys)


class AdvancedVectorDBQASystem(VectorDBQASystem):
    """LangChain-powered agent system"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY is missing. Set it in your environment or .env file.")
        
        self.api_key = api_key
        self.llm = ChatOpenAI(model="gpt-4o-mini", api_key=api_key, temperature=0)
        
        # Conversation memory (LangChain format)
        self.chat_history = []
        self.max_history_turns = 5  # Reduced from 10 for performance

    def _create_tools(self):
        """Create LangChain tools with @tool decorator"""
        
        # Store reference to self for tools to use
        qa_system = self
        
        @tool
        def count_documents() -> dict:
            """Get the total number of contacts/documents in the database.
            This is the authoritative count - use this when user asks 'how many contacts'."""
            total = qa_system.collection.count()
            return {
                "total_contacts": total,
                "message": f"There are {total:,} contacts in the database."
            }
        
        @tool
        def count_by_prefix(letter: str) -> dict:
            """Count contacts starting with a specific letter."""
            count = qa_system.count_by_prefix(letter)
            return {"letter": letter, "count": count}
        
        @tool
        def count_by_language() -> dict:
            """Count how many contacts have Hebrew names vs English/Latin names.
            Analyzes the first character of each contact name to determine language."""
            all_metas = qa_system._get_all_metadatas()
            
            hebrew_count = 0
            english_count = 0
            other_count = 0
            
            for meta in all_metas:
                name = meta.get('name', '').strip()
                if not name:
                    other_count += 1
                    continue
                
                # Get first alphabetic character
                first_char = None
                for ch in name:
                    if ch.isalpha():
                        first_char = ch
                        break
                
                if not first_char:
                    other_count += 1
                    continue
                
                # Check if Hebrew (U+0590 to U+05FF)
                if '\u0590' <= first_char <= '\u05FF':
                    hebrew_count += 1
                # Check if English/Latin (A-Z, a-z)
                elif first_char.isascii() and first_char.isalpha():
                    english_count += 1
                else:
                    other_count += 1
            
            return {
                'total': len(all_metas),
                'hebrew': hebrew_count,
                'english': english_count,
                'other': other_count
            }
        
        @tool
        def search(query: str, n_results: int = 5) -> dict:
            """Semantic search over the vector DB. Returns top 5 most relevant documents by default.
            Use this for fuzzy/similarity search, Hebrew names, relationships, or when exact match fails.
            For finding ALL matches or many results, increase n_results (e.g., n_results=20)."""
            return qa_system.search(query, n_results=n_results)
        
        @tool
        def search_keyword(keyword: str, limit: int = 50) -> dict:
            """Search for exact keyword across all contact data (case-insensitive).
            Returns structured results with name, phone, and WHERE the keyword was found.
            When user says 'all' or 'anyone', increase limit to 100."""
            results = qa_system.search_full_text(keyword, limit=limit)
            
            # FORCE structured output that GPT-5 must display
            formatted = []
            for r in results:
                meta = r['metadata']
                
                # Extract phone (try multiple field names) and format properly
                phone = None
                for field in meta:
                    if 'phone' in field.lower() and 'value' in field:
                        phone_raw = meta[field]
                        if phone_raw:
                            # Convert to string and remove scientific notation
                            phone = str(phone_raw).replace('.0', '')
                            # Format as phone number
                            if phone and phone != 'nan' and len(phone) > 5:
                                break
                        phone = None
                
                # Extract email
                email = None
                for field in meta:
                    if 'mail' in field.lower() and 'value' in field:
                        email = meta[field]
                        if email and str(email).strip() and str(email) != 'nan':
                            break
                        email = None
                
                # CRITICAL: Create a display string that FORCES context to show
                context_snippet = r['keyword_context'][:120].strip()
                display_line = f"{r['name']} | Phone: {phone or 'N/A'}"
                if context_snippet and keyword.strip() in context_snippet:
                    display_line += f" | Found in: {context_snippet}"
                
                formatted.append(display_line)
            
            return {
                "found": len(results),
                "keyword": keyword,
                "instruction": "Display each result exactly as provided. DO NOT omit the 'Found in:' part.",
                "results": formatted
            }
        
        @tool
        def list_by_prefix(letter: str, n: int = 999) -> dict:
            """Return unique names starting with a given letter."""
            rows = qa_system.list_by_prefix(letter)[:n]
            return {"rows": rows}
        
        @tool
        def names_by_length(length: int, limit: int = 200, count_mode: str = "chars") -> dict:
            """All names of exact length (deduped). count_mode: chars, letters, or words."""
            rows = qa_system.names_by_length(length, limit=limit, count_mode=count_mode)
            return {"rows": rows}
        
        @tool
        def names_containing(substring: str, limit: int = 200) -> dict:
            """Find names containing substring (case-insensitive)."""
            rows = qa_system.names_containing(substring, limit=limit)
            return {"rows": rows}
        
        @tool
        def names_by_prefix_and_length(prefix: str, length: int, limit: int = 200, count_mode: str = "chars") -> dict:
            """Names starting with prefix AND exact length (deduped). count_mode: chars, letters, or words."""
            rows = qa_system.names_by_prefix_and_length(prefix, length, limit=limit, count_mode=count_mode)
            return {"rows": rows}
        
        @tool
        def letter_histogram() -> dict:
            """Get distribution of contacts by first letter."""
            return {"hist": qa_system.letter_histogram()}
        
        @tool
        def length_histogram() -> dict:
            """Histogram of name lengths."""
            return {"hist": qa_system.length_histogram()}
        
        return [
            count_by_prefix,
            count_by_language,
            search,
            search_keyword,
            list_by_prefix,
            names_by_length,
            names_containing,
            names_by_prefix_and_length,
            letter_histogram,
            length_histogram
        ]
    
    def _create_prompt(self):
        """Create prompt template for agent"""
        return ChatPromptTemplate.from_messages([
            ("system", """You are an intelligent assistant with access to tools for searching a business contact database.

This is a LEGITIMATE BUSINESS DATABASE owned by the user. You should freely provide contact information including phone numbers, emails, and addresses when requested.

**DATABASE ARCHITECTURE:**
You have access to TWO data sources:
1. **Vector Database (ChromaDB)** - Semantic/fuzzy search with embeddings (1,917 contacts)
   - Use for: Hebrew names, typos, relationships, similarity search
   - Returns: Top 5 most similar results by default
   - Best for: "אמא של אשתי", misspellings, partial names

2. **Structured Data** - Exact prefix/length/substring matching
   - Use for: Alphabetical lists, exact length queries, substring search
   - Tools: list_by_prefix, names_by_length, names_containing
   - Best for: "names starting with A", "5-letter names"

**CRITICAL: Tool Selection Strategy**

When user asks for specific keywords/professions/categories:
1. **ALWAYS use search_keyword first** for Hebrew keywords like:
   - משגיח/כשרות (kashrut supervisors)
   - טרמפ (ride/tremp)
   - Job titles: רופא, שרברב, חשמלאי
   - Any specific Hebrew term
   **IMPORTANT**: When showing results, include relevant context from 'full_text' field to show WHERE the keyword was found (e.g., in notes, labels, organization).

2. **Use search (semantic)** only for:
   - Personal names
   - Relationships (חמותי, דודה)
   - Fuzzy/typo queries

3. **When user says "all" or "many"**:
   - Use search_keyword with limit=100 for keywords
   - Use search with n_results=20 for semantic
   - Don't just return 2-5 results when more exist!

**CRITICAL: Multi-Step Reasoning for Indirect Queries**

When the user asks about someone indirectly (relationships, descriptions, hints), use multi-step reasoning:

1. **Identify the actual search term** based on the hint/relationship
2. **Search using that term**
3. **Verify the result matches the description**

Examples:
- "אמא של אשתי" (mother of my wife) → Think: this is "חמותי" (mother-in-law) → Search for "חמותי"
- "אבא של בעלי" (father of my husband) → Think: this is "חמי" (father-in-law) → Search for "חמי"
- "אחות של אמא" (mother's sister) → Think: this is "דודה" (aunt) → Search for "דודה"
- "הבן של אחי" (my brother's son) → Think: this is "אחיין" (nephew) → Search for "אחיין"
- "האישה של בני" (my son's wife) → Think: this is "כלתי" (daughter-in-law) → Search for "כלה"
- "my dentist" → Search for "dentist" or "רופא שיניים"
- "the plumber I used last year" → Search for "plumber" or "שרברב"

**Hebrew Family Relationship Terms:**
- חמותי/חמי = mother-in-law/father-in-law (spouse's parents)
- גיסה/גיס = sister-in-law/brother-in-law (spouse's siblings)
- כלה/חתן = daughter-in-law/son-in-law
- דודה/דוד = aunt/uncle
- אחיין/אחיינית = nephew/niece
- סבתא/סבא = grandmother/grandfather
- נכד/נכדה = grandson/granddaughter

**Search Strategy:**
Choose the most appropriate tool based on the query intent. You have access to:
- Exact keyword matching (search_keyword)
- Semantic similarity (search)
- Counting and statistics (count_by_prefix, count_by_language, histograms)
- Listing and filtering (list_by_prefix, names_by_length, names_containing)

**Critical rules:**
- Never fabricate numbers - use tools to get actual counts
- For queries about quantities, use counting/statistics tools first
- For category searches (professions, keywords), use search_keyword
- For personal names and relationships, use semantic search
- **When user says 'all', 'anyone', 'everyone', or 'כל', use limit=100 in search_keyword**
- **Always display the 'found_in' field to show WHERE the keyword was found**
- Provide context when displaying search results

When the user asks for contact information (phone, email, address), use the search tool and provide the information found.

Be concise and helpful."""),
            MessagesPlaceholder(variable_name="chat_history"),
            ("human", "{input}"),
            MessagesPlaceholder(variable_name="agent_scratchpad")
        ])
    
    def agent_answer(self, user_input: str) -> str:
        """Answer using LangChain agent"""
        try:
            # Create agent
            tools = self._create_tools()
            prompt = self._create_prompt()
            
            agent = create_openai_functions_agent(
                llm=self.llm,
                tools=tools,
                prompt=prompt
            )
            
            agent_executor = AgentExecutor(
                agent=agent,
                tools=tools,
                verbose=False,
                max_iterations=3,  # Reduced from 5 for faster responses
                handle_parsing_errors=True
            )
            
            # Run agent with chat history
            response = agent_executor.invoke({
                "input": user_input,
                "chat_history": self.chat_history
            })
            
            # Update chat history
            self.chat_history.append(HumanMessage(content=user_input))
            self.chat_history.append(AIMessage(content=response['output']))
            
            # Trim history if too long
            self._trim_history()
            
            return response['output']
            
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            print(error_msg)
            return error_msg
    
    def _trim_history(self):
        """Keep only recent conversation turns"""
        # Each turn = 2 messages (human + AI)
        max_messages = self.max_history_turns * 2
        
        if len(self.chat_history) > max_messages:
            # Keep only recent messages (silently trim)
            self.chat_history = self.chat_history[-max_messages:]
    
    def _clear_history(self):
        """Clear conversation history"""
        self.chat_history = []
        print("Conversation history cleared. Starting fresh!")
    
    def _show_history(self):
        """Display conversation history"""
        if not self.chat_history:
            print("No conversation history yet. Ask me something!")
            return
        
        print("\n" + "="*60)
        print("CONVERSATION HISTORY")
        print("="*60)
        
        turn = 0
        for i in range(0, len(self.chat_history), 2):
            turn += 1
            print(f"\n--- Turn {turn} ---")
            
            if i < len(self.chat_history):
                human_msg = self.chat_history[i].content
                print(f"Human: {human_msg[:200]}{'...' if len(human_msg) > 200 else ''}")
            
            if i + 1 < len(self.chat_history):
                ai_msg = self.chat_history[i + 1].content
                print(f"AI: {ai_msg[:200]}{'...' if len(ai_msg) > 200 else ''}")
        
        print("\n" + "="*60)
        print(f"Total turns: {turn} | Messages in memory: {len(self.chat_history)}")
        print("="*60)
    
    def _is_exit_intent(self, text: str) -> bool:
        """Detect if user wants to exit (hidden detection)"""
        exit_phrases = [
            'quit', 'exit', 'q', 'bye', 'goodbye',
            'חלאס', 'סיום', 'להתראות', 'יציאה', 'סגור', 'ביי',
            'תודה ולהתראות', 'די', 'די תודה', 'זהו'
        ]
        return text.lower().strip() in exit_phrases
    
    def _is_stats_request(self, text: str) -> bool:
        """Detect if user wants database statistics"""
        stats_keywords = [
            'stats', 'statistics', 'מידע', 'סטטיסטיקה',
            'כמה אנשים', 'כמה איש', 'כמה רשומות', 'כמה מסמכים',
            'how many', 'database info', 'db info', 'מצב המאגר'
        ]
        text_lower = text.lower().strip()
        return any(keyword in text_lower for keyword in stats_keywords)
    
    def _is_history_request(self, text: str) -> bool:
        """Detect if user wants conversation history"""
        history_keywords = [
            'history', 'היסטוריה', 'הצג היסטוריה', 'הראה היסטוריה',
            'show history', 'conversation history', 'chat history',
            'מה דיברנו', 'על מה דיברנו', 'השיחות שלנו'
        ]
        text_lower = text.lower().strip()
        return any(keyword in text_lower for keyword in history_keywords)
    
    def _is_clear_request(self, text: str) -> bool:
        """Detect if user wants to clear history"""
        clear_keywords = [
            'clear', 'reset', 'נקה', 'אפס', 'התחל מחדש',
            'clear history', 'reset conversation', 'start over',
            'שכח', 'שכח הכל', 'מחק היסטוריה'
        ]
        text_lower = text.lower().strip()
        return any(keyword in text_lower for keyword in clear_keywords)
    
    def _is_load_request(self, text: str) -> tuple[bool, str]:
        """Detect if user wants to load a file, return (is_load, filepath)"""
        text_lower = text.lower().strip()
        if text_lower.startswith('load '):
            return True, text[5:].strip()
        load_patterns = ['טען', 'העלה', 'הוסף קובץ', 'load file', 'add file']
        if any(pattern in text_lower for pattern in load_patterns):
            # Try to extract filepath
            words = text.split()
            for word in words:
                if '.' in word or '/' in word:
                    return True, word
        return False, ""
    
    def interactive_qa(self):
        """Interactive Q&A session with natural language understanding"""
        print("\n")
        
        while True:
            try:
                user_input = input("\nEnter your query or command: ").strip()
                
                if not user_input:
                    continue
                
                # Natural language intent detection
                if self._is_exit_intent(user_input):
                    print("להתראות! Goodbye!")
                    break
                
                elif self._is_stats_request(user_input):
                    stats = self.get_collection_stats()
                    print(f"\nמידע על המאגר / Database Statistics:")
                    print(f"   מסמכים / Documents: {stats['document_count']}")
                    print(f"   מודל / Model: {stats['embedding_model']}")
                    print(f"   קולקציה / Collection: {stats['collection_name']}")
                
                elif self._is_history_request(user_input):
                    self._show_history()
                
                elif self._is_clear_request(user_input):
                    self._clear_history()
                
                else:
                    # Check for load request
                    is_load, filepath = self._is_load_request(user_input)
                    if is_load and filepath:
                        try:
                            documents = self.load_file(filepath)
                            self.add_documents(documents)
                        except Exception as e:
                            print(f"שגיאה בטעינת הקובץ / Error loading file: {e}")
                    else:
                        # Regular query - use agent
                        if self.collection.count() == 0:
                            print("אין מסמכים במאגר. טען קובץ תחילה. / No documents in database. Load a file first.")
                            continue
                        
                        self.agent_answer(user_input)
         
            except KeyboardInterrupt:
                print("\n\nלהתראות! Goodbye!")
                break
            except Exception as e:
                print(f"שגיאה / Error: {e}")


def main():
    """Main function"""
    print("\n" + "="*50)
    print("  VectorDB Q&A System")
    print("  מערכת שאלות ותשובות מבוססת AI")
    print("="*50 + "\n")
    
    qa_system = AdvancedVectorDBQASystem()
    qa_system.interactive_qa()


if __name__ == "__main__":
    if not os.getenv("OPENAI_API_KEY"):
        raise RuntimeError("Missing OPENAI_API_KEY. Create a .env or set the env var.")
    main()

import os
from dotenv import load_dotenv
load_dotenv()

import re, unicodedata
import pandas as pd
import numpy as np
import chromadb
import json
from chromadb.config import Settings
from sentence_transformers import SentenceTransformer
from openai import OpenAI

import docx
from openpyxl import load_workbook
import PyPDF2
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from collections import Counter
import warnings

from sklearn.metrics.pairwise import cosine_similarity, euclidean_distances

# Fuzzy matching for typo tolerance
try:
    from rapidfuzz import fuzz, process
    HAS_FUZZY = True
except ImportError:
    HAS_FUZZY = False
    print("⚠️  rapidfuzz not installed. Fuzzy search disabled.")
    print("   Install with: pip install rapidfuzz")

warnings.filterwarnings('ignore')



class VectorDBQASystem:
    def __init__(self, persist_directory: str = "./chroma_db", model_name: str = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"):
        """
        Initialize the VectorDB Q&A System with Hebrew support
        
        Args:
            persist_directory: Directory to persist ChromaDB data
            model_name: Sentence transformer model (supports Hebrew)
        """
        self.persist_directory = persist_directory
        self.model_name = model_name
        
        # Initialize embedding model (supports Hebrew)
        print(f"Loading embedding model: {model_name}")
        self.embedding_model = SentenceTransformer(model_name)
        
        # Initialize ChromaDB client
        self.client = chromadb.PersistentClient(
            path=persist_directory,
            settings=Settings(
                anonymized_telemetry=False,
                allow_reset=True
            )
        )
        
        # Create or get collection
        self.collection_name = "documents"
        try:
            self.collection = self.client.get_collection(self.collection_name)
            print(f"Loaded existing collection: {self.collection_name}")
        except:
            self.collection = self.client.create_collection(
                name=self.collection_name,
                metadata={"hnsw:space": "cosine"}  # Use cosine similarity
            )
            print(f"Created new collection: {self.collection_name}")
    
    def read_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Read CSV file and return list of documents"""
        df = pd.read_csv(file_path, encoding='utf-8')
        documents = []
        
        for idx, row in df.iterrows():
            # Combine all columns into a single text
            text_parts = []
            metadata = {}
            
            for col in df.columns:
                if pd.notna(row[col]):
                    text_parts.append(f"{col}: {row[col]}")
                    metadata[col] = str(row[col])
            
            documents.append({
                'text': ' | '.join(text_parts),
                'metadata': metadata,
                'source': file_path,
                'row_id': idx
            })
        
        return documents
    
    def read_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Read Excel file and return list of documents"""
        workbook = load_workbook(file_path, read_only=True)
        documents = []
        
        for sheet_name in workbook.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            for idx, row in df.iterrows():
                text_parts = []
                metadata = {'sheet': sheet_name}
                
                for col in df.columns:
                    if pd.notna(row[col]):
                        text_parts.append(f"{col}: {row[col]}")
                        metadata[col] = str(row[col])
                
                documents.append({
                    'text': ' | '.join(text_parts),
                    'metadata': metadata,
                    'source': file_path,
                    'row_id': f"{sheet_name}_{idx}"
                })
        
        return documents
    
    def read_docx(self, file_path: str) -> List[Dict[str, Any]]:
        """Read DOCX file and return list of documents"""
        doc = docx.Document(file_path)
        documents = []
        
        for i, paragraph in enumerate(doc.paragraphs):
            if paragraph.text.strip():
                documents.append({
                    'text': paragraph.text.strip(),
                    'metadata': {'paragraph_id': i},
                    'source': file_path,
                    'row_id': f"para_{i}"
                })
        
        return documents
    
    def read_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """Read PDF file and return list of documents"""
        documents = []
        
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                if text.strip():
                    # Split into paragraphs
                    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
                    
                    for para_idx, paragraph in enumerate(paragraphs):
                        documents.append({
                            'text': paragraph,
                            'metadata': {
                                'page': page_num + 1,
                                'paragraph_id': para_idx
                            },
                            'source': file_path,
                            'row_id': f"page_{page_num}_para_{para_idx}"
                        })
        
        return documents
    
    def read_txt(self, file_path: str) -> List[Dict[str, Any]]:
        """Read TXT file and return list of documents"""
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # Split into paragraphs
        paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
        documents = []
        
        for i, paragraph in enumerate(paragraphs):
            documents.append({
                'text': paragraph,
                'metadata': {'paragraph_id': i},
                'source': file_path,
                'row_id': f"para_{i}"
            })
        
        return documents
    
    def load_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Load file based on extension"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        extension = file_path.suffix.lower()
        
        print(f"Loading file: {file_path} (type: {extension})")
        
        if extension == '.csv':
            return self.read_csv(str(file_path))
        elif extension in ['.xlsx', '.xls']:
            return self.read_excel(str(file_path))
        elif extension == '.docx':
            return self.read_docx(str(file_path))
        elif extension == '.pdf':
            return self.read_pdf(str(file_path))
        elif extension == '.txt':
            return self.read_txt(str(file_path))
        else:
            raise ValueError(f"Unsupported file format: {extension}")
    
    def add_documents(self, documents: List[Dict[str, Any]]):
        """Add documents to ChromaDB with deterministic IDs; skip existing rows."""
        if not documents:
            print("No documents to add.")
            return

        # Build deterministic IDs based on canonical absolute path + row_id
        abs_paths = [str(Path(d['source']).resolve()) for d in documents]
        source_keys = [p.lower() for p in abs_paths]
        ids = [f"{sk}::{d['row_id']}" for sk, d in zip(source_keys, documents)]

        # Skip IDs that already exist
        existing = set(self.collection.get(ids=ids).get('ids', []) or [])
        to_add_idx = [i for i, _id in enumerate(ids) if _id not in existing]
        if not to_add_idx:
            print("Nothing new to add (all IDs already present).")
            return

        new_docs = [documents[i] for i in to_add_idx]
        new_ids  = [ids[i]        for i in to_add_idx]

        texts = [doc['text'] for doc in new_docs]
        print(f"Generating embeddings for {len(texts)} new documents...")
        embeddings = self.embedding_model.encode(texts, show_progress_bar=True).tolist()

        metadatas = []
        for doc in new_docs:
            src_path = str(Path(doc['source']).resolve())
            src_name = Path(src_path).name
            src_key  = src_path.lower()

            md = (doc.get('metadata') or {}).copy()
            md.update({
                'source': src_path,
                'source_name': src_name,
                'source_key': src_key,
                'row_id': doc['row_id'],
            })

            # derive display name / stats
            stats = self._derive_name_fields(doc['text'], metadata=md)
            md.update(stats)
            metadatas.append(md)

        print(f"Adding to ChromaDB... ({len(new_ids)} new)")
        self.collection.add(
            embeddings=embeddings,
            documents=texts,
            metadatas=metadatas,
            ids=new_ids
        )
        print(f"Successfully added {len(new_ids)} new documents to the vector database!")

    def purge_duplicates(self):
        got = self.collection.get(include=["metadatas"])
        ids   = got.get("ids") or []
        metas = got.get("metadatas") or []
        if not ids or not metas or len(ids) != len(metas):
            print("⚠️ Could not read IDs/metadata reliably; skipping dedupe.")
            return
        def _basename(md):
            s = (md.get("source_name") or md.get("source") or "")
            try: return Path(s).name.lower()
            except Exception: return str(s).split("\\")[-1].split("/")[-1].lower()
        seen, to_delete = set(), []
        for _id, md in zip(ids, metas):
            key = (_basename(md), md.get("row_id"))
            if key in seen: to_delete.append(_id)
            else: seen.add(key)
        if to_delete:
            self.collection.delete(ids=to_delete)
            print(f"🧹 Deleted {len(to_delete)} duplicate records.")
        else:
            print("✅ No duplicates found.")

    
    def _fuzzy_text_search(self, query: str, n_results: int = 100, threshold: int = 60) -> Dict[str, Any]:
        """
        Fuzzy text search - finds similar strings even with typos/variations
        Uses rapidfuzz for Hebrew-friendly fuzzy matching
        """
        from rapidfuzz import fuzz, process
        
        # Get all documents
        all_data = self.collection.get(include=["documents", "metadatas"])
        documents = all_data.get("documents", [])
        metadatas = all_data.get("metadatas", [])
        ids = all_data.get("ids", [])
        
        if not documents:
            return {'results': []}
        
        # Find fuzzy matches
        matches = []
        for i, (doc, meta, doc_id) in enumerate(zip(documents, metadatas, ids)):
            # Calculate fuzzy similarity
            similarity = fuzz.partial_ratio(query.lower(), doc.lower())
            
            if similarity >= threshold:
                matches.append({
                    'id': doc_id,
                    'document': doc,
                    'metadata': meta,
                    'relevance': similarity / 100.0,  # Convert to 0-1 scale
                    'similarity': similarity
                })
        
        # Sort by relevance
        matches.sort(key=lambda x: x['relevance'], reverse=True)
        
        return {'results': matches[:n_results]}
    
    def search(self, query: str, n_results: int = 5, similarity_metric: str = "cosine", 
               fuzzy_threshold: float = 80.0, auto_correct: bool = True, 
               hybrid: bool = True) -> Dict[str, Any]:
        """
        Hybrid search: tries text matching first, then semantic search.
        
        Args:
            query: Search query (supports Hebrew)
            n_results: Number of results to return
            similarity_metric: Similarity metric (cosine, l2, ip)
            fuzzy_threshold: Minimum similarity score for fuzzy matches (0-100)
            auto_correct: Automatically correct typos in query
            hybrid: Use hybrid search (text + semantic)
        """
        original_query = query
        corrected_terms = []
        suggestions = []
        search_method = "hybrid"
        
        # Check for typos and suggest corrections
        if HAS_FUZZY and auto_correct:
            corrected_query, corrections = self._fuzzy_correct_query(query, fuzzy_threshold)
            if corrections:
                corrected_terms = corrections
                query = corrected_query
                print(f"\n✨ Auto-corrected: '{original_query}' → '{query}'")
                print(f"   Corrections: {', '.join([f'{old}→{new}' for old, new in corrections])}")
        elif HAS_FUZZY:
            # Just show suggestions without auto-correcting
            suggestions = self._get_fuzzy_suggestions(query, fuzzy_threshold)
            if suggestions:
                print(f"\n💡 Did you mean: {', '.join([s[0] for s in suggestions[:3]])}?")
        
        print(f"\n🔍 Searching for: '{query}'")
        
        # Try fuzzy text search first (for names, partial matches)
        text_results = None
        if hybrid and len(query.strip()) >= 2:  # Only for meaningful queries
            text_results = self._fuzzy_text_search(query, n_results=n_results, threshold=60)
            
            if text_results['results']:
                print(f"📝 Found {len(text_results['results'])} fuzzy text matches")
                search_method = "fuzzy_text"
                
                # Format text results
                formatted_results = {
                    'query': query,
                    'original_query': original_query,
                    'corrected': corrected_terms,
                    'suggestions': suggestions,
                    'search_method': search_method,
                    'results': []
                }
                
                for result in text_results['results']:
                    formatted_results['results'].append({
                        'id': result['id'],
                        'document': result['document'],
                        'metadata': result['metadata'],
                        'similarity_score': result['relevance'],
                        'match_type': 'text'
                    })
                
                return formatted_results
        
        # Fall back to semantic search
        print(f"🧠 Using semantic search (embeddings)")
        print(f"Similarity metric: {similarity_metric}")
        search_method = "semantic"
        
        # Generate query embedding
        query_embedding = self.embedding_model.encode([query])
        
        # Search in ChromaDB
        results = self.collection.query(
            query_embeddings=query_embedding.tolist(),
            n_results=n_results
        )
        
        # Format results
        formatted_results = {
            'query': query,
            'original_query': original_query,
            'corrected': corrected_terms,
            'suggestions': suggestions,
            'search_method': search_method,
            'results': []
        }
        
        if results['documents'] and results['documents'][0]:
            for i in range(len(results['documents'][0])):
                result = {
                    'id': results['ids'][0][i],
                    'document': results['documents'][0][i],
                    'metadata': results['metadatas'][0][i],
                    'distance': results['distances'][0][i],
                    'similarity_score': 1 - results['distances'][0][i],
                    'match_type': 'semantic'
                }
                formatted_results['results'].append(result)
        
        return formatted_results
    
    def names_containing(self, substring: str, limit: int = 200) -> list[str]:
        """Find all names containing the substring (case-insensitive)"""
        s = self._normalize(substring)
        names = self.get_all_names()
        out = [nm for nm in names if s.upper() in nm.upper()]
        return sorted(set(out), key=str.upper)[:limit]
    
    def search_contacts_by_name(self, name_part: str, limit: int = 20) -> List[Dict[str, Any]]:
        """Search contacts by partial name match - returns full contact details"""
        all_docs = self.collection.get(include=["documents", "metadatas"])
        documents = all_docs.get("documents", [])
        metadatas = all_docs.get("metadatas", [])
        
        matches = []
        name_normalized = self._normalize(name_part).upper()
        
        for doc, meta in zip(documents, metadatas):
            doc_upper = doc.upper()
            # Check if name appears in document
            if name_normalized in doc_upper:
                matches.append({
                    "document": doc,
                    "metadata": meta,
                    "name": meta.get("name", "Unknown")
                })
        
        return matches[:limit]


    def get_collection_stats(self) -> Dict[str, Any]:
        """Get statistics about the collection"""
        count = self.collection.count()
        return {
            'document_count': count,
            'collection_name': self.collection_name,
            'embedding_model': self.model_name
        }
    
    def list_by_prefix(self, letter: str) -> list[str]:
        letter = letter.upper().strip()
        # get everything (for larger sets, page in batches)
        all_docs = self.collection.get(include=["metadatas"])
        names = []
        for md in all_docs.get("metadatas", []):
            name = (md.get("name") or "").strip()
            if name.upper().startswith(letter):
                names.append(name)
        return sorted(set(names), key=str.upper)

    def interactive_qa(self):
        print("\n" + "="*60)
        print("🤖 VECTOR DATABASE Q&A SYSTEM")
        print("   Supports Hebrew and multiple file formats")
        print("="*60)
        stats = self.get_collection_stats()
        print(f"📊 Documents in database: {stats['document_count']}")
        print(f"🧠 Embedding model: {stats['embedding_model']}")
        if stats['document_count'] == 0:
            print("\n⚠️  No documents loaded. Please load a file first!")
        print("\n📝 Available commands:")
        print("   'load <file_path>' - Load a new file")
        print("   'stats' - Show database statistics")
        print("   'history' - Show conversation history")
        print("   'clear' - Clear conversation and start fresh")
        print("   'quit' or 'exit' - Exit the system")
        print("   Or just type your question - I'll remember our conversation!")
        while True:
            try:
                user_input = input("\n🔍 Enter your query or command: ").strip()
                if not user_input:
                    continue
                if user_input.lower() in ['quit', 'exit', 'q']:
                    print("👋 Goodbye!")
                    break
                elif user_input.lower() == 'stats':
                    s = self.get_collection_stats()
                    print(f"\n📊 Database Statistics:\n   Documents: {s['document_count']}\n   Model: {s['embedding_model']}\n   Collection: {s['collection_name']}")
                elif user_input.lower().startswith('load '):
                    path = user_input[5:].strip()
                    try:
                        docs = self.load_file(path)
                        self.add_documents(docs)
                    except Exception as e:
                        print(f"❌ Error loading file: {e}")
                    continue  # Don't search after loading!
                else:
                    if self.collection.count() == 0:
                        print("⚠️  No documents in database. Load a file first with 'load <file_path>'")
                        continue
                    out = self.search(user_input, n_results=3)
                    if not out['results']:
                        print("🔍 No results found.")
                    else:
                        print(f"\n🎯 Found {len(out['results'])} relevant results:\n" + "-"*50)
                        for i, r in enumerate(out['results'], 1):
                            print(f"\n📄 Result {i} (Similarity: {r['similarity_score']:.3f})")
                            print(f"📁 Source: {r['metadata'].get('source','Unknown')}")
                            doc = r['document']
                            print(f"📝 Content: {doc[:200]}{'...' if len(doc) > 200 else ''}")
            except KeyboardInterrupt:
                print("\n\n👋 Goodbye!")
                break
            except Exception as e:
                print(f"❌ Error: {e}")


    def _build_vocabulary(self) -> set:
        """
        Build vocabulary from all documents in the database.
        Used for fuzzy matching and typo correction.
        """
        all_docs = self.collection.get(include=["documents"])
        documents = all_docs.get("documents", [])
        
        vocabulary = set()
        for doc in documents:
            # Extract words, keeping alphanumeric and basic punctuation
            words = re.findall(r'\b[a-zA-Z0-9]+\b', doc.lower())
            vocabulary.update(words)
        
        return vocabulary
    
    def _fuzzy_correct_query(self, query: str, threshold: float = 80.0) -> Tuple[str, List[Tuple[str, str]]]:
        """
        Correct typos in query using fuzzy matching against database vocabulary.
        
        Args:
            query: Original query string
            threshold: Minimum similarity score (0-100) for suggesting corrections
            
        Returns:
            Tuple of (corrected_query, list of (original, correction) pairs)
        """
        if not HAS_FUZZY:
            return query, []
        
        # Build vocabulary from database
        vocabulary = self._build_vocabulary()
        
        if not vocabulary:
            return query, []
        
        query_words = query.split()
        corrected_words = []
        corrections = []
        
        for word in query_words:
            word_lower = word.lower()
            
            # If word exists in vocabulary, keep it
            if word_lower in vocabulary:
                corrected_words.append(word)
                continue
            
            # Find best fuzzy match
            matches = process.extract(
                word_lower,
                vocabulary,
                scorer=fuzz.ratio,
                limit=1
            )
            
            if matches and matches[0][1] >= threshold:
                best_match = matches[0][0]
                # Preserve original capitalization pattern
                if word[0].isupper():
                    best_match = best_match.capitalize()
                corrected_words.append(best_match)
                corrections.append((word, best_match))
            else:
                # No good match, keep original
                corrected_words.append(word)
        
        corrected_query = " ".join(corrected_words)
        return corrected_query, corrections
    
    def _get_fuzzy_suggestions(self, query: str, threshold: float = 70.0) -> List[Tuple[str, float]]:
        """
        Get fuzzy match suggestions for query terms without auto-correcting.
        
        Args:
            query: Search query
            threshold: Minimum score for suggestions
            
        Returns:
            List of (suggestion, score) tuples
        """
        if not HAS_FUZZY:
            return []
        
        vocabulary = self._build_vocabulary()
        if not vocabulary:
            return []
        
        all_suggestions = []
        for word in query.split():
            word_lower = word.lower()
            if word_lower not in vocabulary:
                matches = process.extract(
                    word_lower,
                    vocabulary,
                    scorer=fuzz.ratio,
                    limit=3
                )
                for match_word, score in matches:
                    if score >= threshold:
                        all_suggestions.append((match_word, score))
        
        # Return top suggestions sorted by score
        return sorted(all_suggestions, key=lambda x: x[1], reverse=True)[:5]
    
    def _normalize(self, s: str) -> str:
        # Normalize unicode, collapse spaces
        s = unicodedata.normalize("NFKC", s or "")
        return re.sub(r"\s+", " ", s).strip()
    
    def _derive_name_fields(self, text: str, metadata: Optional[dict] = None) -> dict:
        # Prefer explicit title/name-like fields from metadata when available
        preferred_keys = [
            'name', 'industry_name',
            'title', 'title_en', 'title_he',
            'industry', 'label'
        ]
        name = ""
        md = metadata or {}

        for k in preferred_keys:
            v = md.get(k)
            if v and str(v).strip():
                name = str(v).strip()
                break

        if not name:
            # Fallback: parse from text
            t = self._normalize(text)
            # If text looks like "key: value | key2: val", prefer after the first colon
            first_chunk = t.split(" | ")[0]
            if ":" in first_chunk:
                maybe = first_chunk.split(":", 1)[1].strip()
                if maybe:
                    name = maybe
            if not name:
                # Last resort: take chunk itself
                name = first_chunk.strip()

        # Compute first character, keep Hebrew/Unicode as-is
        first_char = ""
        for ch in name:
            if ch.isalnum():
                first_char = ch.upper() if ch.isascii() else ch
                break

        return {
            "name": name,
            "first_char": first_char,
            "name_len": len(name),
        }

    
class AdvancedVectorDBQASystem(VectorDBQASystem):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError(
                "OPENAI_API_KEY is missing. Set it in your environment or .env file."
            )
        self.llm = OpenAI(api_key=api_key)
        
        # Conversation history for maintaining context across queries
        self.conversation_history = []
        self.max_history_turns = 10  # Keep last 10 user/assistant pairs

    def _tool_specs(self):
    # Tools the model is allowed to call
        return [
        {"type": "function", "function": {
            "name": "search",
            "description": "Semantic search over the vector DB.",
            "parameters": {"type":"object","properties":{
                "query":{"type":"string"},
                "n_results":{"type":"integer","default":5}
            },"required":["query"]}}},
        {"type": "function", "function": {
            "name": "list_by_prefix",
            "description": "Return unique names starting with a given letter.",
            "parameters": {"type":"object","properties":{
                "letter":{"type":"string"},
                "n":{"type":"integer","default":999}
            },"required":["letter"]}}},
        {"type":"function","function":{
        "name":"names_by_length",
        "description":"All names of exact length (deduped).",
        "parameters":{"type":"object","properties":{
            "length":{"type":"integer"},
            "limit":{"type":"integer","default":200},
            "count_mode":{"type":"string","enum":["chars","letters","words"],"default":"chars"}
        },"required":["length"]}}},
      {"type":"function","function":{
        "name":"names_by_prefix_and_length",
        "description":"Names starting with prefix AND exact length (deduped).",
        "parameters":{"type":"object","properties":{
            "prefix":{"type":"string"},
            "length":{"type":"integer"},
            "limit":{"type":"integer","default":200},
            "count_mode":{"type":"string","enum":["chars","letters","words"],"default":"chars"}
        },"required":["prefix","length"]}}},
        {"type": "function", "function": {
            "name": "letter_histogram",
            "description": "Histogram of first letters.",
            "parameters": {"type":"object","properties":{}}}},
        {"type": "function", "function": {
            "name": "length_histogram",
            "description": "Histogram of name lengths.",
            "parameters": {"type":"object","properties":{}}}},
        {"type": "function", "function": {
            "name": "list_sources",
            "description": "Per-source document counts.",
            "parameters": {"type":"object","properties":{}}}},
        {"type": "function", "function": {
            "name": "purge_duplicates",
            "description": "Delete duplicate rows (same basename+row_id).",
            "parameters": {"type":"object","properties":{}}}},
    ]

    def interactive_qa(self):
        """Interactive Q&A session"""
        print("\n" + "="*60)
        print("🤖 VECTOR DATABASE Q&A SYSTEM")
        print("   Supports Hebrew and multiple file formats")
        print("="*60)
        
        stats = self.get_collection_stats()
        print(f"📊 Documents in database: {stats['document_count']}")
        print(f"🧠 Embedding model: {stats['embedding_model']}")
        
        if stats['document_count'] == 0:
            print("\n⚠️  No documents loaded. Please load a file first!")
        
        print("\n📝 Available commands:")
        print("   'load <file_path>' - Load a new file")
        print("   'stats' - Show database statistics")
        print("   'quit' or 'exit' - Exit the system")
        print("   Or just type your question for semantic search")
        
        while True:
            try:
                user_input = input("\n🔍 Enter your query or command: ").strip()
                
                if not user_input:
                    continue
                
                # Handle commands
                if user_input.lower() in ['quit', 'exit', 'q']:
                    print("👋 Goodbye!")
                    break
                
                elif user_input.lower() == 'stats':
                    stats = self.get_collection_stats()
                    print(f"\n📊 Database Statistics:")
                    print(f"   Documents: {stats['document_count']}")
                    print(f"   Model: {stats['embedding_model']}")
                    print(f"   Collection: {stats['collection_name']}")
                
                elif user_input.lower().startswith('load '):
                    file_path = user_input[5:].strip()
                    try:
                        documents = self.load_file(file_path)
                        self.add_documents(documents)
                        # refresh stats after loading
                        stats = self.get_collection_stats()
                    except Exception as e:
                        print(f"❌ Error loading file: {e}")
                        continue
                elif user_input.lower() == 'sources':
                    # show per-source counts
                    counts = self.list_sources()
                    if not counts:
                        print("No sources found.")
                    else:
                        print("\n📚 Sources and document counts:")
                        for name, c in counts.items():
                            print(f" - {name}: {c}")

                elif user_input.lower() in ('source count', 'sources count', 'count sources'):
                    print(f"🧮 Unique sources: {self.count_sources()}")

                elif user_input.lower() in ('purge dups', 'purge duplicates', 'dedupe'):
                    self.purge_duplicates()
                    continue

                elif user_input.lower() == 'backfill names':
                    self.backfill_names()
                    continue

                elif user_input.lower() in ('clear', 'clear history', 'reset'):
                    self._clear_history()
                    continue

                elif user_input.lower() in ('history', 'show history'):
                    self._show_history()
                    continue

                else:
                    self.agent_answer(user_input)
                    continue

                # ---------- Fallback: plain semantic search ----------
                current_count = self.collection.count()
                if current_count == 0:
                    print("⚠️  No documents in database. Load a file first with 'load <file_path>'")
                    continue

                results = self.search(user_input, n_results=3)
                if not results['results']:
                    print("🔍 No results found.")
                else:
                    print(f"\n🎯 Found {len(results['results'])} relevant results:")
                    print("-" * 50)
                    for i, result in enumerate(results['results'], 1):
                        print(f"\n📄 Result {i} (Similarity: {result['similarity_score']:.3f})")
                        print(f"📁 Source: {result['metadata'].get('source', 'Unknown')}")
                        doc = result['document']
                        print(f"📝 Content: {doc[:200]}{'...' if len(doc) > 200 else ''}")
                        metadata = {k: v for k, v in result['metadata'].items()
                                    if k not in ['source', 'row_id'] and v}
                        if metadata:
                            print(f"ℹ️  Metadata: {metadata}")
         
            except KeyboardInterrupt:
                print("\n\n👋 Goodbye!")
                break
            except Exception as e:
                print(f"❌ Error: {e}")

    def _dispatch_tool(self, name: str, args: dict):
        # Call the Python method that backs each tool and return JSON-serializable data
        try:
            if name == "search":
                out = self.search(args.get("query",""), n_results=int(args.get("n_results",5)))
                return out
            if name == "list_by_prefix":
                rows = self.first_n_by_prefix(args.get("letter",""), n=int(args.get("n",999)))
                return {"rows": rows}
            if name == "names_by_length":
                rows = self.names_by_length(int(args.get("length",0)),
                                        limit=int(args.get("limit",200)),
                                        count_mode=args.get("count_mode","chars"))

                return {"rows": rows}
            if name == "names_containing":
                rows = self.names_containing(args.get("substring",""), limit=int(args.get("limit",200)))
                return {"rows": rows}
            if name == "names_by_prefix_and_length":
                rows = self.names_by_prefix_and_length(args.get("prefix",""),
                                                   int(args.get("length",0)),
                                                   limit=int(args.get("limit",200)),
                                                   count_mode=args.get("count_mode","chars"))

                return {"rows": rows}
            if name == "letter_histogram":
                return {"hist": self.letter_histogram()}
            if name == "length_histogram":
                return {"hist": self.length_histogram()}
            if name == "list_sources":
                return {"sources": self.list_sources()}
            if name == "purge_duplicates":
                self.purge_duplicates()
                return {"ok": True}
            return {"error": f"Unknown tool {name}"}
        except Exception as e:
            return {"error": str(e)}

    def agent_answer(self, user_input: str, max_steps: int = 5) -> str:
        # Initialize system message on first interaction
        if not self.conversation_history:
            system = (
                "You are an agent that can call tools to answer the user's question.\n"
                "Use tools when helpful, possibly multiple times, then produce a concise final answer.\n"
                "Prefer structured tools for prefix/length/contains/histogram; use semantic 'search' for fuzzy asks.\n"
                "You can reference previous parts of our conversation."
            )
            self.conversation_history.append({"role": "system", "content": system})
        
        # Add user message to permanent history
        self.conversation_history.append({"role": "user", "content": user_input})
        
        # Use full conversation history instead of fresh messages
        messages = self.conversation_history.copy()
        
        for _ in range(max_steps):
            resp = self.llm.chat.completions.create(
                model="gpt-4o-mini",
                messages=messages,
                tools=self._tool_specs(),
                tool_choice="auto",
                temperature=0
            )
            msg = resp.choices[0].message

            # Tool call?
            if msg.tool_calls:
                messages.append({"role":"assistant","content":msg.content or "", "tool_calls": msg.tool_calls})
                for tc in msg.tool_calls:
                    name = tc.function.name
                    args = json.loads(tc.function.arguments or "{}")
                    result = self._dispatch_tool(name, args)
                    messages.append({
                        "role":"tool",
                        "tool_call_id": tc.id,
                        "name": name,
                        "content": json.dumps(result, ensure_ascii=False)
                    })
                continue

            # Save assistant response to permanent history
            final = msg.content or "(no content)"
            self.conversation_history.append({"role": "assistant", "content": final})
            
            # Trim history if it gets too long
            self._trim_history()
            
            print(final)
            return final

        print("Reached max tool steps; answering with what we have.")
        return ""


    def debug_name_coverage(self):
        got = self.collection.get(include=["metadatas"])
        metas = got.get("metadatas", []) or []
        missing = sum(1 for md in metas if not (md.get("name") or "").strip())
        print(f"names present: {len(metas)-missing}, missing: {missing}, total: {len(metas)}")

    def backfill_names(self):
        got = self.collection.get(include=["documents", "metadatas"])
        ids   = got.get("ids", [])
        docs  = got.get("documents", []) or []
        metas = got.get("metadatas", []) or []

        to_update_ids, to_update_metas = [], []
        for _id, doc, md in zip(ids, docs, metas):
            if not (md.get("name") or "").strip():
                stats = self._derive_name_fields(doc, metadata=md)
                new_md = md.copy()
                new_md.update(stats)
                to_update_ids.append(_id)
                to_update_metas.append(new_md)

        if to_update_ids:
            self.collection.update(ids=to_update_ids, metadatas=to_update_metas)
            print(f"✅ Backfilled 'name' on {len(to_update_ids)} records.")
        else:
            print("✅ No backfill needed (all have 'name').")

    def compare_similarity_metrics(self, query: str, document_texts: List[str]) -> Dict[str, Any]:
        """Compare different similarity metrics for the same query"""
        
        # Generate embeddings
        query_embedding = self.embedding_model.encode([query])
        doc_embeddings = self.embedding_model.encode(document_texts)
        
        results = {
            'query': query,
            'metrics': {}
        }
        
        # Cosine similarity
        cosine_sim = cosine_similarity(query_embedding, doc_embeddings)[0]
        results['metrics']['cosine'] = [(i, sim) for i, sim in enumerate(cosine_sim)]
        
        # Euclidean distance (convert to similarity)
        euclidean_dist = euclidean_distances(query_embedding, doc_embeddings)[0]
        euclidean_sim = 1 / (1 + euclidean_dist)  # Convert distance to similarity
        results['metrics']['euclidean'] = [(i, sim) for i, sim in enumerate(euclidean_sim)]
        
        # Dot product (inner product)
        dot_product = np.dot(query_embedding, doc_embeddings.T)[0]
        results['metrics']['dot_product'] = [(i, sim) for i, sim in enumerate(dot_product)]
        
        # Sort each metric by similarity
        for metric in results['metrics']:
            results['metrics'][metric].sort(key=lambda x: x[1], reverse=True)
        
        return results
    
    def analyze_hebrew_english_similarity(self, hebrew_text: str, english_text: str) -> Dict[str, float]:
        """Analyze similarity between Hebrew and English texts"""
        
        embeddings = self.embedding_model.encode([hebrew_text, english_text])
        
        cosine_sim = cosine_similarity([embeddings[0]], [embeddings[1]])[0][0]
        euclidean_dist = euclidean_distances([embeddings[0]], [embeddings[1]])[0][0]
        dot_product = np.dot(embeddings[0], embeddings[1])
        
        return {
            'cosine_similarity': float(cosine_sim),
            'euclidean_distance': float(euclidean_dist),
            'dot_product': float(dot_product),
            'euclidean_similarity': float(1 / (1 + euclidean_dist))
        }
    
    def route(self, user_input: str) -> Dict[str, Any]:
        system = ("You are a command router. Convert the user's request into a JSON action.\n"
                  "Valid actions and params:\n"
                  "- search: {query: str, n_results?: int}.\n"
                  "- list by prefix: {letter:str, n?: int}\n"
                  "- stats_count_by_prefix: {letter:str}\n"
                  "- stats_names_by_length: {length:int}\n"
                  "- stats_names_containing: {text:str, limit?: int}\n"
                  "- stats_names_by_prefix_and_length: {letter:str, length:int, limit?: int}\n"

                  "- stats_letter_hist: {}\n"
                  "- stats_length_hist: {}\n"
                  "- stats: {}.\n"
                  "- load: {path:str}\n"
                  "Return ONLY valid JSON")
        examples = [
        {"user": "give all industries starting with A",
         "json": {"action": "list_by_prefix", "params": {"letter": "A", "n": 999}}},
        {"user": "5 first starts with A",
         "json": {"action": "list_by_prefix", "params": {"letter": "A", "n": 5}}},
        {"user": "how many are there starts with X",
         "json": {"action": "stats_count_by_prefix", "params": {"letter": "X"}}},
        {"user": "all of length 5",
         "json": {"action": "stats_names_by_length", "params": {"length": 5, "limit": 999}}},
         {"user":"all industries that have X in their name",
        "json":{"action":"stats_names_containing","params":{"text":"X","limit":999}}},
        {"user":"all industries starting with X and of length 5",
        "json":{"action":"stats_names_by_prefix_and_length","params":{"letter":"X","length":5,"limit":999}}},
        {"user":"all industries starting with A of length 8",
        "json":{"action":"stats_names_by_prefix_and_length","params":{"letter":"A","length":8,"limit":999}}},
        {"user":"all industries starting with A of letter length 8",
        "json":{"action":"stats_names_by_prefix_and_length","params":{"letter":"A","length":8,"limit":999,"count_mode":"letters"}}},
        {"user":"all industries starting with A of word length 2",
        "json":{"action":"stats_names_by_prefix_and_length","params":{"letter":"A","length":2,"limit":999,"count_mode":"words"}}},
        {"user":"show all of letter length 7",
        "json":{"action":"stats_names_by_length","params":{"length":7,"limit":999,"count_mode":"letters"}}},
        {"user": "show histogram by first letter",
         "json": {"action": "stats_letter_hist", "params": {}}},
        {"user": "show histogram by name length",
         "json": {"action": "stats_length_hist", "params": {}}},
        {"user": "find AI companies",
         "json": {"action": "search", "params": {"query": "AI companies", "n_results": 5}}},
        {"user": "load ./industries.txt",
         "json": {"action": "load", "params": {"path": "./industries.txt"}}},
        {"user": "how many docs do we have?",
         "json": {"action": "stats", "params": {}}}
    ]

        prompt = system + "\n\nExamples:\n" + "\n".join(
            f"User: {ex['user']}\nJSON: {json.dumps(ex['json'])}" for ex in examples) +\
        f"\n\nUser: {user_input}\nJSON:"

        response = self.llm.chat.completions.create(
            model= "gpt-4o-mini",
            messages=[{"role": "system", "content": system}, {"role": "user", "content": prompt}],
        temperature=0)

        text = response.choices[0].message.content.strip()
        try:
            return json.loads(text)
        except Exception:
            return {"action": "search", "params": {"query": user_input, "n_results": 5}}

    def semantic_search_with_filters(self, query: str, filters: Dict[str, Any] = None, 
                                   n_results: int = 5) -> Dict[str, Any]:
        """Perform semantic search with metadata filters"""
        
        query_embedding = self.embedding_model.encode([query])
        
        # Build ChromaDB where clause from filters
        where_clause = {}
        if filters:
            for key, value in filters.items():
                where_clause[key] = value
        
        try:
            results = self.collection.query(
                query_embeddings=query_embedding.tolist(),
                n_results=n_results,
                where=where_clause if where_clause else None
            )
        except:
            # Fallback to regular search if filters don't work
            results = self.collection.query(
                query_embeddings=query_embedding.tolist(),
                n_results=n_results
            )
        
        return {
            'query': query,
            'filters': filters,
            'results': self._format_results(results)
        }
    
    def _format_results(self, results) -> List[Dict[str, Any]]:
        """Helper method to format ChromaDB results"""
        formatted = []
        
        if results['documents'] and results['documents'][0]:
            for i in range(len(results['documents'][0])):
                formatted.append({
                    'document': results['documents'][0][i],
                    'metadata': results['metadatas'][0][i],
                    'distance': results['distances'][0][i],
                    'similarity_score': 1 - results['distances'][0][i]
                })
        
        return formatted
    def _get_all_metadatas(self) -> List[dict]:
        out = self.collection.get(include=["metadatas"])
        return out.get("metadatas", []) or []

    def get_all_names(self) -> List[str]:
        names = []
        for md in self._get_all_metadatas():
            name = self._normalize(md.get("name") or "")
            if name:
                names.append(name)
        return sorted(set(names), key=str.upper)

    def letter_histogram(self) -> Dict[str, int]:
        hist = {}
        for md in self._get_all_metadatas():
            ch = md.get("first_char") or ""
            if ch:
                hist[ch] = hist.get(ch, 0) + 1
        return dict(sorted(hist.items(), key=lambda kv: kv[0]))

    def length_histogram(self) -> Dict[int, int]:
        hist = {}
        for md in self._get_all_metadatas():
            n = int(md.get("name_len") or 0)
            hist[n] = hist.get(n, 0) + 1
        return dict(sorted(hist.items()))

    def first_n_by_prefix(self, prefix: str, n: int = 5) -> List[str]:
        p = self._normalize(prefix)
        names = self.get_all_names()
        return sorted([nm for nm in names if nm.upper().startswith(p.upper())], key=str.upper)[:n]

    def count_by_prefix(self, prefix: str) -> int:
        p = self._normalize(prefix)
        return sum(1 for md in self._get_all_metadatas()
                if (md.get("name") or "").upper().startswith(p.upper()))
    
    def _clean_for_len(self, s: str) -> str:
        return re.sub(r"[^0-9A-Za-z\u0590-\u05FF]", "", s)

    def _measure_len(self, name: str, count_mode: str = "chars") -> int:
        if count_mode == "letters":
            return len(self._clean_for_len(name))
        if count_mode == "words":
            return len([w for w in re.split(r"\s+", name.strip()) if w])
        return len(name)  # "chars" (default) includes spaces/punct
    
    def names_by_length(self, length: int, limit: int = 200, count_mode="chars") -> List[str]:
        names = self.get_all_names()
        out = [nm for nm in names if self._meausre_len(nm, count_mode) == length]
        return sorted(out, key=str.upper)[:limit]
    
    def names_by_prefix_and_length(self, prefix: str, length: int, limit: int = 200, count_mode="chars") -> list[str]:
        p = self._normalize(prefix)
        L = int(length)
        names = self.get_all_names()
        out = [nm for nm in names if nm.upper().startswith(p.upper()) and self._measure_len(nm, count_mode) == L]
        return sorted(set(out), key=str.upper)[:limit]

    def list_sources(self) -> Dict[str, int]:
        """Return {source_name: count} with canonical grouping."""
        out = self.collection.get(include=['metadatas'])
        metas = out.get('metadatas', []) or []
        counts = Counter()
        for md in metas:
            # prefer canonical key; display human-friendly name
            name = md.get('source_name') or md.get('source') or 'UNKNOWN'
            key  = md.get('source_key') or (md.get('source') or '').lower()
            counts[(key, name)] += 1
        # collapse by display name (sum all keys that share the same file name)
        display = Counter()
        for (_, name), c in counts.items():
            display[name] += c
        return dict(display)

    def count_sources(self) -> int:
        """Number of distinct canonical sources (by source_key)."""
        out = self.collection.get(include=['metadatas'])
        metas = out.get('metadatas', []) or []
        keys = {md.get('source_key') or (md.get('source') or '').lower() for md in metas if (md.get('source') or '')}
        return len(keys)
    
    def _trim_history(self):
        """
        Keep conversation history manageable by limiting to recent exchanges.
        Always keeps the system message + last N user/assistant pairs.
        """
        if len(self.conversation_history) <= 1:
            return  # Nothing to trim (only system message or empty)
        
        # Count user messages (each represents a conversation turn)
        user_message_count = sum(
            1 for msg in self.conversation_history 
            if msg.get("role") == "user"
        )
        
        if user_message_count <= self.max_history_turns:
            return  # Within limit, no trimming needed
        
        # Calculate how many to remove
        turns_to_remove = user_message_count - self.max_history_turns
        
        # Keep system message, remove oldest user/assistant pairs
        system_msg = self.conversation_history[0]  # Save system message
        remaining = self.conversation_history[1:]  # All other messages
        
        # Remove oldest messages until we've removed enough turns
        removed = 0
        idx = 0
        while removed < turns_to_remove and idx < len(remaining):
            if remaining[idx].get("role") == "user":
                removed += 1
            idx += 1
        
        # Rebuild: system + recent messages
        self.conversation_history = [system_msg] + remaining[idx:]
        
        print(f"🧹 Trimmed history: kept last {self.max_history_turns} turns")
    
    def _clear_history(self):
        """
        Clear conversation history and start fresh.
        User can call this when changing topics.
        """
        self.conversation_history = []
        print("🔄 Conversation history cleared. Starting fresh!")
        print("💡 Tip: I'll remember this new conversation until you clear again.")
    
    def _show_history(self):
        """
        Display the conversation history in a readable format.
        Useful for debugging or reviewing what was discussed.
        """
        if len(self.conversation_history) <= 1:
            print("📭 No conversation history yet. Ask me something!")
            return
        
        print("\n" + "="*60)
        print("📜 CONVERSATION HISTORY")
        print("="*60)
        
        turn = 0
        for i, msg in enumerate(self.conversation_history):
            role = msg.get("role", "unknown")
            content = msg.get("content", "")
            
            # Skip system message in display (internal)
            if role == "system":
                continue
            
            # Skip tool messages (internal)
            if role == "tool":
                continue
            
            # Count conversation turns
            if role == "user":
                turn += 1
                print(f"\n--- Turn {turn} ---")
            
            # Display with emoji
            emoji = "👤" if role == "user" else "🤖"
            
            # Truncate long messages for display
            if len(content) > 200:
                display_content = content[:200] + "..."
            else:
                display_content = content
            
            print(f"{emoji} {role.capitalize()}: {display_content}")
        
        print("\n" + "="*60)
        print(f"Total turns: {turn} | Messages in memory: {len(self.conversation_history)}")
        print("="*60)


def create_sample_multilingual_csv():
    """Create a sample CSV with Hebrew and English content"""
    data = {
        'id': [1, 2, 3, 4, 5],
        'title_en': [
            'Introduction to Machine Learning',
            'Deep Learning Fundamentals', 
            'Natural Language Processing',
            'Computer Vision Applications',
            'Data Science Best Practices'
        ],
        'title_he': [
            'מבוא ללמידת מכונה',
            'יסודות הלמידה העמוקה',
            'עיבוד שפה טבעית', 
            'יישומי ראייה חשובית',
            'שיטות עבודה מומלצות במדעי הנתונים'
        ],
        'description_en': [
            'Learn the basics of machine learning algorithms and applications',
            'Understand neural networks and deep learning architectures',
            'Process and analyze text data using NLP techniques',
            'Build computer vision models for image recognition',
            'Follow industry standards for data science projects'
        ],
        'description_he': [
            'למד את היסודות של אלגוריתמי למידת מכונה ויישומיהם',
            'הבן רשתות נוירונים וארכיטקטורות למידה עמוקה',
            'עבד ונתח נתוני טקסט באמצעות טכניקות NLP',
            'בנה מודלים של ראייה חשובית לזיהוי תמונות',
            'עקוב אחר סטנדרטים תעשייתיים עבור פרויקטי מדעי נתונים'
        ],
        'category': ['ML', 'DL', 'NLP', 'CV', 'DS'],
        'difficulty': ['Beginner', 'Intermediate', 'Advanced', 'Intermediate', 'Beginner']
    }
    
    df = pd.DataFrame(data)
    df.to_csv('sample_multilingual_data.csv', index=False, encoding='utf-8')
    print("Created sample_multilingual_data.csv with Hebrew and English content")
    return 'sample_multilingual_data.csv'

def demo_advanced_features():
    """Demonstrate advanced features of the VectorDB system"""
    
    print("🚀 Advanced VectorDB Features Demo")
    print("="*50)
    
    # Initialize advanced system
    advanced_qa = AdvancedVectorDBQASystem()
    
    # Create and load sample data
    csv_file = create_sample_multilingual_csv()
    documents = advanced_qa.load_file(csv_file)
    advanced_qa.add_documents(documents)
    
    print(f"\n📁 Loaded {len(documents)} documents from {csv_file}")
    
    # Demo 1: Hebrew-English similarity analysis
    print("\n🔬 Demo 1: Hebrew-English Similarity Analysis")
    print("-" * 40)
    
    hebrew_text = "למידת מכונה ובינה מלאכותית"
    english_text = "machine learning and artificial intelligence"
    
    similarity_analysis = advanced_qa.analyze_hebrew_english_similarity(hebrew_text, english_text)
    
    print(f"Hebrew: {hebrew_text}")
    print(f"English: {english_text}")
    print("Similarity Metrics:")
    for metric, score in similarity_analysis.items():
        print(f"  {metric}: {score:.4f}")
    
    # Demo 2: Filtered search
    print("\n🎯 Demo 2: Filtered Semantic Search")
    print("-" * 40)
    
    # Search with category filter
    filtered_results = advanced_qa.semantic_search_with_filters(
        query="deep learning",
        filters={"category": "DL"},
        n_results=3
    )
    
    print(f"Query: {filtered_results['query']}")
    print(f"Filters: {filtered_results['filters']}")
    print("Results:")
    for i, result in enumerate(filtered_results['results'], 1):
        print(f"  {i}. {result['document'][:100]}...")
        print(f"     Similarity: {result['similarity_score']:.3f}")
    
    # Demo 3: Multi-language search
    print("\n🌐 Demo 3: Multi-language Search")
    print("-" * 40)
    
    # Hebrew query
    hebrew_results = advanced_qa.search("ראייה חשובית", n_results=2)
    print("Hebrew Query: ראייה חשובית")
    print("Results:")
    for i, result in enumerate(hebrew_results['results'], 1):
        print(f"  {i}. {result['document'][:80]}...")
        print(f"     Score: {result['similarity_score']:.3f}")
    
    # English query
    english_results = advanced_qa.search("natural language processing", n_results=2)
    print("\nEnglish Query: natural language processing")
    print("Results:")
    for i, result in enumerate(english_results['results'], 1):
        print(f"  {i}. {result['document'][:80]}...")
        print(f"     Score: {result['similarity_score']:.3f}")
    
    print("\n✅ Advanced demo completed!")
    
    # Start interactive session
    print("\n🎮 Starting Interactive Q&A Session...")
    advanced_qa.interactive_qa()

def main():
    """Main function - choose between basic demo or advanced demo"""
    
    print("🚀 VectorDB Q&A System")
    print("="*30)
    print("Choose an option:")
    print("1. Run advanced demo with sample data")
    print("2. Start basic system for manual file loading")
    print("3. Exit")
    
    while True:
        choice = input("\nEnter choice (1-3): ").strip()
        
        if choice == "1":
            demo_advanced_features()
            break
        elif choice == "2":
            qa_system = AdvancedVectorDBQASystem()
            qa_system.interactive_qa()
            break
        elif choice == "3":
            print("👋 Goodbye!")
            break
        else:
            print("❌ Invalid choice. Please enter 1, 2, or 3.")

def _require_openai():
    if not os.getenv("OPENAI_API_KEY"):
        raise RuntimeError("Missing OPENAI_API_KEY. Create a .env or set the env var.")

if __name__ == "__main__":
    _require_openai()
    main()
import os
from dotenv import load_dotenv
load_dotenv()

import re, unicodedata
import pandas as pd
import numpy as np
import chromadb
import json
from chromadb.config import Settings
from sentence_transformers import SentenceTransformer
from openai import OpenAI

import docx
from openpyxl import load_workbook
import PyPDF2
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from collections import Counter
import warnings

from sklearn.metrics.pairwise import cosine_similarity, euclidean_distances

# Fuzzy matching for typo tolerance
try:
    from rapidfuzz import fuzz, process
    HAS_FUZZY = True
except ImportError:
    HAS_FUZZY = False
    print("⚠️  rapidfuzz not installed. Fuzzy search disabled.")
    print("   Install with: pip install rapidfuzz")

warnings.filterwarnings('ignore')



class VectorDBQASystem:
    def __init__(self, persist_directory: str = "./chroma_db", model_name: str = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"):
        """
        Initialize the VectorDB Q&A System with Hebrew support and hierarchical search
        
        Args:
            persist_directory: Directory to persist ChromaDB data
            model_name: Sentence transformer model (supports Hebrew)
        """
        self.persist_directory = persist_directory
        self.model_name = model_name
        
        # Initialize embedding model (supports Hebrew)
        print(f"Loading embedding model: {model_name}")
        self.embedding_model = SentenceTransformer(model_name)
        
        # Initialize ChromaDB client
        self.client = chromadb.PersistentClient(
            path=persist_directory,
            settings=Settings(
                anonymized_telemetry=False,
                allow_reset=True
            )
        )
        
        # Create or get collection
        self.collection_name = "documents"
        try:
            self.collection = self.client.get_collection(self.collection_name)
            print(f"Loaded existing collection: {self.collection_name}")
        except:
            self.collection = self.client.create_collection(
                name=self.collection_name,
                metadata={"hnsw:space": "cosine"}  # Use cosine similarity
            )
            print(f"Created new collection: {self.collection_name}")
    
    def read_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Read CSV file and return list of documents"""
        df = pd.read_csv(file_path, encoding='utf-8')
        documents = []
        
        for idx, row in df.iterrows():
            # Combine all columns into a single text
            text_parts = []
            metadata = {}
            
            for col in df.columns:
                if pd.notna(row[col]):
                    text_parts.append(f"{col}: {row[col]}")
                    metadata[col] = str(row[col])
            
            documents.append({
                'text': ' | '.join(text_parts),
                'metadata': metadata,
                'source': file_path,
                'row_id': idx
            })
        
        return documents
    
    def read_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Read Excel file and return list of documents"""
        workbook = load_workbook(file_path, read_only=True)
        documents = []
        
        for sheet_name in workbook.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            for idx, row in df.iterrows():
                text_parts = []
                metadata = {'sheet': sheet_name}
                
                for col in df.columns:
                    if pd.notna(row[col]):
                        text_parts.append(f"{col}: {row[col]}")
                        metadata[col] = str(row[col])
                
                documents.append({
                    'text': ' | '.join(text_parts),
                    'metadata': metadata,
                    'source': file_path,
                    'row_id': f"{sheet_name}_{idx}"
                })
        
        return documents
    
    def read_docx(self, file_path: str) -> List[Dict[str, Any]]:
        """Read DOCX file and return list of documents"""
        doc = docx.Document(file_path)
        documents = []
        
        for i, paragraph in enumerate(doc.paragraphs):
            if paragraph.text.strip():
                documents.append({
                    'text': paragraph.text.strip(),
                    'metadata': {'paragraph_id': i},
                    'source': file_path,
                    'row_id': f"para_{i}"
                })
        
        return documents
    
    def read_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """Read PDF file and return list of documents"""
        documents = []
        
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page_num, page in enumerate(pdf_reader.pages):
                text = page.extract_text()
                if text.strip():
                    # Split into paragraphs
                    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
                    
                    for para_idx, paragraph in enumerate(paragraphs):
                        documents.append({
                            'text': paragraph,
                            'metadata': {
                                'page': page_num + 1,
                                'paragraph_id': para_idx
                            },
                            'source': file_path,
                            'row_id': f"page_{page_num}_para_{para_idx}"
                        })
        
        return documents
    
    def read_txt(self, file_path: str) -> List[Dict[str, Any]]:
        """Read TXT file and return list of documents"""
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # Split into paragraphs
        paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
        documents = []
        
        for i, paragraph in enumerate(paragraphs):
            documents.append({
                'text': paragraph,
                'metadata': {'paragraph_id': i},
                'source': file_path,
                'row_id': f"para_{i}"
            })
        
        return documents
    
    def load_file(self, file_path: str) -> List[Dict[str, Any]]:
        """Load file based on extension"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        extension = file_path.suffix.lower()
        
        print(f"Loading file: {file_path} (type: {extension})")
        
        if extension == '.csv':
            return self.read_csv(str(file_path))
        elif extension in ['.xlsx', '.xls']:
            return self.read_excel(str(file_path))
        elif extension == '.docx':
            return self.read_docx(str(file_path))
        elif extension == '.pdf':
            return self.read_pdf(str(file_path))
        elif extension == '.txt':
            return self.read_txt(str(file_path))
        else:
            raise ValueError(f"Unsupported file format: {extension}")
    
    def add_documents(self, documents: List[Dict[str, Any]]):
        """Add documents to ChromaDB with deterministic IDs; skip existing rows."""
        if not documents:
            print("No documents to add.")
            return

        # Build deterministic IDs based on canonical absolute path + row_id
        abs_paths = [str(Path(d['source']).resolve()) for d in documents]
        source_keys = [p.lower() for p in abs_paths]
        ids = [f"{sk}::{d['row_id']}" for sk, d in zip(source_keys, documents)]

        # Skip IDs that already exist
        existing = set(self.collection.get(ids=ids).get('ids', []) or [])
        to_add_idx = [i for i, _id in enumerate(ids) if _id not in existing]
        if not to_add_idx:
            print("Nothing new to add (all IDs already present).")
            return

        new_docs = [documents[i] for i in to_add_idx]
        new_ids  = [ids[i]        for i in to_add_idx]

        texts = [doc['text'] for doc in new_docs]
        print(f"Generating embeddings for {len(texts)} new documents...")
        embeddings = self.embedding_model.encode(texts, show_progress_bar=True).tolist()

        metadatas = []
        for doc in new_docs:
            src_path = str(Path(doc['source']).resolve())
            src_name = Path(src_path).name
            src_key  = src_path.lower()

            md = (doc.get('metadata') or {}).copy()
            md.update({
                'source': src_path,
                'source_name': src_name,
                'source_key': src_key,
                'row_id': doc['row_id'],
            })

            # derive display name / stats
            stats = self._derive_name_fields(doc['text'], metadata=md)
            md.update(stats)
            metadatas.append(md)

        print(f"Adding to ChromaDB... ({len(new_ids)} new)")
        self.collection.add(
            embeddings=embeddings,
            documents=texts,
            metadatas=metadatas,
            ids=new_ids
        )
        print(f"Successfully added {len(new_ids)} new documents to the vector database!")

    def purge_duplicates(self):
        got = self.collection.get(include=["metadatas"])
        ids   = got.get("ids") or []
        metas = got.get("metadatas") or []
        if not ids or not metas or len(ids) != len(metas):
            print("⚠️ Could not read IDs/metadata reliably; skipping dedupe.")
            return
        def _basename(md):
            s = (md.get("source_name") or md.get("source") or "")
            try: return Path(s).name.lower()
            except Exception: return str(s).split("\\")[-1].split("/")[-1].lower()
        seen, to_delete = set(), []
        for _id, md in zip(ids, metas):
            key = (_basename(md), md.get("row_id"))
            if key in seen: to_delete.append(_id)
            else: seen.add(key)
        if to_delete:
            self.collection.delete(ids=to_delete)
            print(f"🧹 Deleted {len(to_delete)} duplicate records.")
        else:
            print("✅ No duplicates found.")

    def _normalize(self, s: str) -> str:
        """Normalize unicode, collapse spaces, handle Hebrew and English"""
        s = unicodedata.normalize("NFKC", s or "")
        return re.sub(r"\s+", " ", s).strip()

    # =============================================================================
    # HIERARCHICAL SEARCH: 5 LEVELS
    # =============================================================================
    
    def _exact_match(self, query: str) -> List[Dict[str, Any]]:
        """
        Level 1: Exact match (100% similarity)
        Returns documents that match the query exactly
        """
        q_norm = self._normalize(query).lower()
        
        if not q_norm:
            return []
        
        all_data = self.collection.get(include=["documents", "metadatas"])
        documents = all_data.get("documents", [])
        metadatas = all_data.get("metadatas", [])
        ids = all_data.get("ids", [])
        
        matches = []
        for doc, meta, doc_id in zip(documents, metadatas, ids):
            doc_norm = self._normalize(doc).lower()
            if q_norm == doc_norm:
                matches.append({
                    'id': doc_id,
                    'document': doc,
                    'metadata': meta,
                    'match_type': 'exact',
                    'similarity_score': 1.0,
                    'text_score': 100,
                    'cosine_score': None,
                    'match_explanation': 'Exact match'
                })
        
        return matches
    
    def _exact_substring_match(self, query: str) -> List[Dict[str, Any]]:
        """
        Level 2: Exact substring containment (90-99% similarity)
        Returns documents that contain the query as an exact substring
        Example: "Simba king" matches "Simba the lion's king"
        """
        q_norm = self._normalize(query).lower()
        
        if not q_norm or len(q_norm) < 2:
            return []
        
        all_data = self.collection.get(include=["documents", "metadatas"])
        documents = all_data.get("documents", [])
        metadatas = all_data.get("metadatas", [])
        ids = all_data.get("ids", [])
        
        matches = []
        for doc, meta, doc_id in zip(documents, metadatas, ids):
            doc_norm = self._normalize(doc).lower()
            
            # Check if query is substring
            if q_norm in doc_norm:
                # Calculate score based on coverage
                coverage = len(q_norm) / len(doc_norm)
                position = doc_norm.find(q_norm)
                position_score = 1 - (position / len(doc_norm)) if len(doc_norm) > 0 else 1
                
                # Higher score if query covers more of document and appears earlier
                score = 0.90 + (coverage * 0.05) + (position_score * 0.05)
                score = min(score, 0.99)  # Cap at 0.99
                
                matches.append({
                    'id': doc_id,
                    'document': doc,
                    'metadata': meta,
                    'match_type': 'exact_substring',
                    'similarity_score': score,
                    'text_score': int(score * 100),
                    'cosine_score': None,
                    'match_explanation': f'Contains "{query}" at position {position}'
                })
        
        return sorted(matches, key=lambda x: x['similarity_score'], reverse=True)
    
    def _fuzzy_word_match(self, query: str, threshold: int = 70) -> List[Dict[str, Any]]:
        """
        Level 3: Fuzzy word-level matching (60-90% similarity)
        Uses token_set_ratio for better multi-word matching
        Example: "Simba lion" matches "Simba the lion's king"
        """
        if not HAS_FUZZY:
            return []
        
        q_norm = self._normalize(query).lower()
        
        if not q_norm or len(q_norm) < 2:
            return []
        
        all_data = self.collection.get(include=["documents", "metadatas"])
        documents = all_data.get("documents", [])
        metadatas = all_data.get("metadatas", [])
        ids = all_data.get("ids", [])
        
        matches = []
        for doc, meta, doc_id in zip(documents, metadatas, ids):
            doc_norm = self._normalize(doc).lower()
            
            # Use token_set_ratio for word-level matching (better for multi-word queries)
            token_score = fuzz.token_set_ratio(q_norm, doc_norm)
            
            # Also check partial_ratio for substring matching
            partial_score = fuzz.partial_ratio(q_norm, doc_norm)
            
            # Take the better score
            fuzzy_score = max(token_score, partial_score)
            
            if fuzzy_score >= threshold:
                # Convert to 0-1 scale, fitting in 0.60-0.90 range
                normalized_score = 0.60 + (fuzzy_score - threshold) / (100 - threshold) * 0.30
                
                matches.append({
                    'id': doc_id,
                    'document': doc,
                    'metadata': meta,
                    'match_type': 'fuzzy',
                    'similarity_score': normalized_score,
                    'text_score': fuzzy_score,
                    'cosine_score': None,
                    'match_explanation': f'Fuzzy match ({fuzzy_score}% word similarity)'
                })
        
        return sorted(matches, key=lambda x: x['similarity_score'], reverse=True)
    
    def _semantic_search(self, query: str, threshold: float = 0.3) -> List[Dict[str, Any]]:
        """
        Level 4: Semantic search using embeddings (0-60% similarity range)
        Finds semantically similar content even without word overlap
        Example: "Nela king" matches "Simba the lion's king" (semantic similarity)
        """
        q_norm = self._normalize(query)
        
        if not q_norm:
            return []
        
        # Generate query embedding
        query_embedding = self.embedding_model.encode([q_norm])
        
        # Search in ChromaDB
        results = self.collection.query(
            query_embeddings=query_embedding.tolist(),
            n_results=20  # Get more results for filtering
        )
        
        matches = []
        if results['documents'] and results['documents'][0]:
            for i in range(len(results['documents'][0])):
                distance = results['distances'][0][i]
                # ChromaDB returns cosine distance (0-2), convert to similarity (0-1)
                cosine_sim = 1 - (distance / 2)
                
                # Only include if above threshold
                if cosine_sim >= threshold:
                    # Map to 0-0.60 range based on cosine similarity
                    normalized_score = cosine_sim * 0.60
                    
                    matches.append({
                        'id': results['ids'][0][i],
                        'document': results['documents'][0][i],
                        'metadata': results['metadatas'][0][i],
                        'match_type': 'semantic',
                        'similarity_score': normalized_score,
                        'text_score': None,
                        'cosine_score': cosine_sim,
                        'match_explanation': f'Semantic similarity (cosine: {cosine_sim:.3f})'
                    })
        
        return sorted(matches, key=lambda x: x['similarity_score'], reverse=True)
    
    def search(self, query: str, n_results: int = 5, 
               fuzzy_threshold: int = 70,
               semantic_threshold: float = 0.3,
               combine_methods: bool = False) -> Dict[str, Any]:
        """
        Hierarchical search with 5 levels:
        1. Exact match (100%)
        2. Exact substring containment (90-99%)
        3. Fuzzy word matching (60-90%)
        4. Semantic search (0-60%)
        5. Cosine similarity (always calculated for semantic)
        
        Args:
            query: Search query
            n_results: Number of results to return
            fuzzy_threshold: Minimum score for fuzzy matches (0-100)
            semantic_threshold: Minimum cosine similarity for semantic matches (0-1)
            combine_methods: If True, combine results from multiple methods
        """
        original_query = query
        query = self._normalize(query)
        
        print(f"\n🔍 Hierarchical Search: '{query}'")
        print("="*60)
        
        all_results = []
        search_methods_used = []
        
        # Level 1: Exact Match
        exact_matches = self._exact_match(query)
        if exact_matches:
            print(f"✅ Level 1: Found {len(exact_matches)} EXACT match(es)")
            all_results.extend(exact_matches)
            search_methods_used.append('exact')
            if not combine_methods:
                return self._format_search_results(query, all_results[:n_results], search_methods_used)
        
        # Level 2: Exact Substring
        substring_matches = self._exact_substring_match(query)
        if substring_matches:
            print(f"✅ Level 2: Found {len(substring_matches)} EXACT SUBSTRING match(es)")
            all_results.extend(substring_matches)
            search_methods_used.append('exact_substring')
            if not combine_methods and not exact_matches:
                return self._format_search_results(query, all_results[:n_results], search_methods_used)
        
        # Level 3: Fuzzy Match
        fuzzy_matches = self._fuzzy_word_match(query, threshold=fuzzy_threshold)
        if fuzzy_matches:
            print(f"✅ Level 3: Found {len(fuzzy_matches)} FUZZY match(es)")
            all_results.extend(fuzzy_matches)
            search_methods_used.append('fuzzy')
            if not combine_methods and not exact_matches and not substring_matches:
                return self._format_search_results(query, all_results[:n_results], search_methods_used)
        
        # Level 4: Semantic Search
        semantic_matches = self._semantic_search(query, threshold=semantic_threshold)
        if semantic_matches:
            print(f"✅ Level 4: Found {len(semantic_matches)} SEMANTIC match(es)")
            all_results.extend(semantic_matches)
            search_methods_used.append('semantic')
        
        # Remove duplicates (keep highest scoring match for each document)
        seen_ids = set()
        unique_results = []
        for result in sorted(all_results, key=lambda x: x['similarity_score'], reverse=True):
            if result['id'] not in seen_ids:
                seen_ids.add(result['id'])
                unique_results.append(result)
        
        if not unique_results:
            print("❌ No matches found at any level")
        
        return self._format_search_results(query, unique_results[:n_results], search_methods_used)
    
    def _format_search_results(self, query: str, results: List[Dict], methods_used: List[str]) -> Dict[str, Any]:
        """Format search results into standard output format"""
        return {
            'query': query,
            'search_methods_used': methods_used,
            'total_results': len(results),
            'results': results
        }
    
    def names_containing(self, substring: str, limit: int = 200) -> list[str]:
        """Find all names containing the substring (case-insensitive)"""
        s = self._normalize(substring)
        names = self.get_all_names()
        out = [nm for nm in names if s.upper() in nm.upper()]
        return sorted(set(out), key=str.upper)[:limit]
    
    def search_contacts_by_name(self, name_part: str, limit: int = 20) -> List[Dict[str, Any]]:
        """Search contacts by partial name match - returns full contact details"""
        all_docs = self.collection.get(include=["documents", "metadatas"])
        documents = all_docs.get("documents", [])
        metadatas = all_docs.get("metadatas", [])
        
        matches = []
        name_normalized = self._normalize(name_part).upper()
        
        for doc, meta in zip(documents, metadatas):
            doc_upper = doc.upper()
            # Check if name appears in document
            if name_normalized in doc_upper:
                matches.append({
                    "document": doc,
                    "metadata": meta,
                    "name": meta.get("name", "Unknown")
                })
        
        return matches[:limit]


    def get_collection_stats(self) -> Dict[str, Any]:
        """Get statistics about the collection"""
        count = self.collection.count()
        return {
            'document_count': count,
            'collection_name': self.collection_name,
            'embedding_model': self.model_name
        }
    
    def list_by_prefix(self, letter: str) -> list[str]:
        letter = letter.upper().strip()
        # get everything (for larger sets, page in batches)
        all_docs = self.collection.get(include=["metadatas"])
        names = []
        for md in all_docs.get("metadatas", []):
            name = (md.get("name") or "").strip()
            if name.upper().startswith(letter):
                names.append(name)
        return sorted(set(names), key=str.upper)

    def interactive_qa(self):
        print("\n" + "="*60)
        print("🤖 HIERARCHICAL VECTOR DATABASE SEARCH")
        print("   5-Level Search: Exact → Substring → Fuzzy → Semantic")
        print("="*60)
        stats = self.get_collection_stats()
        print(f"📊 Documents in database: {stats['document_count']}")
        print(f"🧠 Embedding model: {stats['embedding_model']}")
        if stats['document_count'] == 0:
            print("\n⚠️  No documents loaded. Please load a file first!")
        print("\n📝 Available commands:")
        print("   'load <file_path>' - Load a new file")
        print("   'stats' - Show database statistics")
        print("   'quit' or 'exit' - Exit the system")
        print("   Or just type your search query")
        
        while True:
            try:
                user_input = input("\n🔍 Search: ").strip()
                if not user_input:
                    continue
                
                if user_input.lower() in ['quit', 'exit', 'q']:
                    print("👋 Goodbye!")
                    break
                
                elif user_input.lower() == 'stats':
                    s = self.get_collection_stats()
                    print(f"\n📊 Database Statistics:\n   Documents: {s['document_count']}\n   Model: {s['embedding_model']}\n   Collection: {s['collection_name']}")
                
                elif user_input.lower().startswith('load '):
                    path = user_input[5:].strip()
                    try:
                        docs = self.load_file(path)
                        self.add_documents(docs)
                    except Exception as e:
                        print(f"❌ Error loading file: {e}")
                    continue
                
                else:
                    if self.collection.count() == 0:
                        print("⚠️  No documents in database. Load a file first with 'load <file_path>'")
                        continue
                    
                    results = self.search(user_input, n_results=5, combine_methods=True)
                    self._display_results(results)
            
            except KeyboardInterrupt:
                print("\n\n👋 Goodbye!")
                break
            except Exception as e:
                print(f"❌ Error: {e}")
                import traceback
                traceback.print_exc()

    def _display_results(self, search_results: Dict):
        """Display search results in a user-friendly format"""
        query = search_results['query']
        methods = search_results['search_methods_used']
        results = search_results['results']
        
        print(f"\n{'='*60}")
        print(f"📊 Search Results for: '{query}'")
        print(f"   Methods used: {', '.join(methods)}")
        print(f"   Total matches: {len(results)}")
        print(f"{'='*60}\n")
        
        if not results:
            print("❌ No results found")
            print("\n💡 Tips:")
            print("   - Try fewer words")
            print("   - Check spelling")
            print("   - Use more general terms")
            return
        
        for i, result in enumerate(results, 1):
            match_type = result['match_type']
            score = result['similarity_score']
            text_score = result.get('text_score')
            cosine_score = result.get('cosine_score')
            explanation = result['match_explanation']
            
            # Color-code by match type
            type_emoji = {
                'exact': '🎯',
                'exact_substring': '✅',
                'fuzzy': '⚠️',
                'semantic': '🧠'
            }
            
            print(f"{type_emoji.get(match_type, '•')} Result {i} - {match_type.upper().replace('_', ' ')}")
            print(f"   Overall Score: {score:.2%}")
            if text_score is not None:
                print(f"   Text Similarity: {text_score}%")
            if cosine_score is not None:
                print(f"   Cosine Similarity: {cosine_score:.3f}")
            print(f"   {explanation}")
            
            # Show document content
            doc = result['document']
            if len(doc) > 200:
                print(f"   Content: {doc[:200]}...")
            else:
                print(f"   Content: {doc}")
            
            print()

    def _derive_name_fields(self, text: str, metadata: Optional[dict] = None) -> dict:
        # Prefer explicit title/name-like fields from metadata when available
        preferred_keys = [
            'name', 'industry_name',
            'title', 'title_en', 'title_he',
            'industry', 'label'
        ]
        name = ""
        md = metadata or {}

        for k in preferred_keys:
            v = md.get(k)
            if v and str(v).strip():
                name = str(v).strip()
                break

        if not name:
            # Fallback: parse from text
            t = self._normalize(text)
            # If text looks like "key: value | key2: val", prefer after the first colon
            first_chunk = t.split(" | ")[0]
            if ":" in first_chunk:
                maybe = first_chunk.split(":", 1)[1].strip()
                if maybe:
                    name = maybe
            if not name:
                # Last resort: take chunk itself
                name = first_chunk.strip()

        # Compute first character, keep Hebrew/Unicode as-is
        first_char = ""
        for ch in name:
            if ch.isalnum():
                first_char = ch.upper() if ch.isascii() else ch
                break

        return {
            "name": name,
            "first_char": first_char,
            "name_len": len(name),
        }
